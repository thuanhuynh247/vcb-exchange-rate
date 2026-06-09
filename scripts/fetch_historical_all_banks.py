"""
Fetch and backfill USD/VND exchange rates for 6 banks from 2026-01-01 to present.
Updates D:\\Tygia-Tudong\\TyGia_Banking.xlsx.

API endpoints discovered:
- VCB: https://www.vietcombank.com.vn/api/exchangerates?date=YYYY-MM-DD
- BIDV: https://bidv.com.vn/ServicesBIDV/ExchangeDetailServlet?date=DD/MM/YYYY
- VietinBank: POST https://www.vietinbank.vn/ty-gia-khcn (Server Action)
  - History: next-action ff24b60505a8da357a655878afe7dd2d1f9f0e52, payload ["YYYY-MM-DD","YYYY-MM-DD","USD","transfer_rate|sell_rate"]
  - Current: next-action 1e43a43a5124d6cc3cb463bc54021b34f39a4065, payload ["YYYY-MM-DDT15:45:00","USD"]
- Techcombank: GET https://techcombank.com/content/techcombank/web/vn/vi/cong-cu-tien-ich/ty-gia/_jcr_content.exchange-rates.integration.json (today only)
- ACB: https://acb.com.vn/api/front/v1/currency?currency=VND&effectiveDateTime=YYYY-MM-DDT14:00:00.000
- SeaBank: POST https://www.seabank.com.vn/cong-cu-tien-ich/ty-gia (Server Action)
  - next-action d65f2411081b93638167328d79ca76cf2bc7ec18, payload ["DD/MM/YYYY"]
"""
import os
os.makedirs(r"D:\Tygia-Tudong\Temp", exist_ok=True)
os.environ["TEMP"] = r"D:\Tygia-Tudong\Temp"
os.environ["TMP"] = r"D:\Tygia-Tudong\Temp"
os.environ["OPENPYXL_LXML"] = "False"
import openpyxl
openpyxl.xml.LXML = False
import requests
import pandas as pd
from datetime import datetime, timedelta
import os
import sys
import re
import json
import urllib3
import concurrent.futures
from openpyxl import load_workbook

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Config
OUTPUT_FILE = r'D:\Tygia-Tudong\TyGia_Banking.xlsx'
START_DATE = datetime(2026, 1, 1)
END_DATE = datetime.now()

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36'
}

BANKS = ["Vietcombank", "VietinBank", "BIDV", "Techcombank", "ACB", "SeaBank"]

# =============================================================================
# VietinBank history cache (fetched once for all dates)
# =============================================================================
_vietinbank_transfer_cache = {}  # key: "YYYY-MM-DD" -> {"buy_transfer": ..., "sell": ...}
_vietinbank_sell_cache = {}

def _load_vietinbank_history():
    """Fetch all VietinBank history in 2 requests (transfer_rate + sell_rate)."""
    global _vietinbank_transfer_cache, _vietinbank_sell_cache
    if _vietinbank_transfer_cache:
        return
    
    url = "https://www.vietinbank.vn/ty-gia-khcn"
    action_history = "ff24b60505a8da357a655878afe7dd2d1f9f0e52"
    h = {**HEADERS, "Content-Type": "text/plain;charset=UTF-8",
         "next-action": action_history, "accept": "text/x-component",
         "referer": "https://www.vietinbank.vn/ty-gia-khcn"}
    
    start_str = START_DATE.strftime('%Y-%m-%d')
    end_str = END_DATE.strftime('%Y-%m-%d')
    
    try:
        # Transfer rate history
        r = requests.post(url, headers=h, data=json.dumps([start_str, end_str, "USD", "transfer_rate"]), 
                         verify=False, timeout=15)
        for line in r.text.strip().split('\n'):
            if line.startswith("1:"):
                data = json.loads(line[2:])
                for item in data:
                    date_key = item.get('apply_date', '')
                    _vietinbank_transfer_cache[date_key] = item.get('close', item.get('open'))
        print(f"  VietinBank transfer_rate history: {len(_vietinbank_transfer_cache)} records")
        
        # Sell rate history
        r = requests.post(url, headers=h, data=json.dumps([start_str, end_str, "USD", "sell_rate"]),
                         verify=False, timeout=15)
        for line in r.text.strip().split('\n'):
            if line.startswith("1:"):
                data = json.loads(line[2:])
                for item in data:
                    date_key = item.get('apply_date', '')
                    _vietinbank_sell_cache[date_key] = item.get('close', item.get('open'))
        print(f"  VietinBank sell_rate history: {len(_vietinbank_sell_cache)} records")
    except Exception as e:
        print(f"  ⚠ Lỗi fetch VietinBank history: {e}")


def fetch_vietinbank_usd_current(date_str_yyyy_mm_dd):
    """Fetch VietinBank current rate via Server Action (for today's data)."""
    url = "https://www.vietinbank.vn/ty-gia-khcn"
    action = "1e43a43a5124d6cc3cb463bc54021b34f39a4065"
    h = {**HEADERS, "Content-Type": "text/plain;charset=UTF-8",
         "next-action": action, "accept": "text/x-component",
         "referer": "https://www.vietinbank.vn/ty-gia-khcn"}
    
    try:
        r = requests.post(url, headers=h, data=json.dumps([f"{date_str_yyyy_mm_dd}T15:45:00", "USD"]),
                         verify=False, timeout=10)
        for line in r.text.strip().split('\n'):
            if line.startswith("1:"):
                data = json.loads(line[2:])
                if data and isinstance(data, list) and len(data) > 0:
                    item = data[0]
                    return {
                        'buy_cash': item.get('cash_rate_big'),
                        'buy_transfer': item.get('transfer_rate'),
                        'sell': item.get('sell_rate')
                    }
    except Exception as e:
        pass
    return None


def get_vietinbank_usd_for_date(date_str_yyyy_mm_dd):
    """Get VietinBank USD rate from cache or fetch current."""
    _load_vietinbank_history()
    
    transfer_rate = _vietinbank_transfer_cache.get(date_str_yyyy_mm_dd)
    sell_rate = _vietinbank_sell_cache.get(date_str_yyyy_mm_dd)
    
    if transfer_rate or sell_rate:
        return {
            'buy_cash': transfer_rate,  # VietinBank: cash_buy ~ transfer_buy
            'buy_transfer': transfer_rate,
            'sell': sell_rate
        }
    return None


# =============================================================================
# SeaBank history cache
# =============================================================================
_seabank_cache = {}  # key: "DD/MM/YYYY" -> rates

def get_seabank_usd_for_date(date_str_dd_mm_yyyy):
    """Fetch SeaBank USD rate for a specific date via Server Action."""
    if date_str_dd_mm_yyyy in _seabank_cache:
        return _seabank_cache[date_str_dd_mm_yyyy]
    
    url = "https://www.seabank.com.vn/cong-cu-tien-ich/ty-gia"
    action = "d65f2411081b93638167328d79ca76cf2bc7ec18"
    h = {**HEADERS, "Content-Type": "text/plain;charset=UTF-8",
         "next-action": action, "accept": "text/x-component",
         "referer": "https://www.seabank.com.vn/cong-cu-tien-ich/ty-gia"}
    
    try:
        r = requests.post(url, headers=h, data=json.dumps([date_str_dd_mm_yyyy]),
                         verify=False, timeout=10)
        for line in r.text.strip().split('\n'):
            if line.startswith("1:"):
                data = json.loads(line[2:])
                if data and isinstance(data, dict) and 'details' in data:
                    usd = next((x for x in data['details'] if x.get('currency') == 'USD'), None)
                    if usd:
                        result = {
                            'buy_cash': _to_float(usd.get('buy')),
                            'buy_transfer': _to_float(usd.get('transferBuy')),
                            'sell': _to_float(usd.get('sell'))
                        }
                        _seabank_cache[date_str_dd_mm_yyyy] = result
                        return result
    except Exception as e:
        pass
    return None


def fetch_techcombank_usd_historical(date_str_yyyy_mm_dd):
    """Fetch historical Techcombank USD rate for a specific date (YYYY-MM-DD)."""
    url = f"https://techcombank.com/content/techcombank/web/vn/vi/cong-cu-tien-ich/ty-gia/_jcr_content.exchange-rates.{date_str_yyyy_mm_dd}.integration.json"
    try:
        r = requests.get(url, headers=HEADERS, verify=False, timeout=10)
        if r.status_code == 200:
            data = r.json()
            rates = data.get("exchangeRate", {}).get("data", [])
            for item in rates:
                if item.get("label", "").startswith("USD (50") or item.get("label") == "USD (50,100)":
                    return {
                        'buy_cash': clean_rate_val(item.get('bidRateTM')),
                        'buy_transfer': clean_rate_val(item.get('bidRateCK')),
                        'sell': clean_rate_val(item.get('askRate'))
                    }
    except Exception as e:
        pass
    return None


def fetch_techcombank_usd_current():
    """Fetch current Techcombank USD rate (today only - no historical API).
    Uses session to initialize CSRF token first (required for API to return data).
    """
    csrf_url = "https://techcombank.com/libs/granite/csrf/token.json"
    rate_url = "https://techcombank.com/content/techcombank/web/vn/vi/cong-cu-tien-ich/ty-gia/_jcr_content.exchange-rates.integration.json"
    session_headers = {**HEADERS,
                       "Accept": "application/json, text/plain, */*",
                       "Referer": "https://techcombank.com/cong-cu-tien-ich/ty-gia"}
    try:
        session = requests.Session()
        # Initialize session with CSRF call (sets cookies needed for rate endpoint)
        session.get(csrf_url, headers=session_headers, verify=False, timeout=10)
        r = session.get(rate_url, headers=session_headers, verify=False, timeout=10)
        if r.status_code == 200:
            data = r.json()
            for item in data.get("exchangeRate", {}).get("data", []):
                if item.get("label", "").startswith("USD (50"):  # USD (50,100) large bills
                    return {
                        'buy_cash': _to_float(item.get('bidRateTM')),
                        'buy_transfer': _to_float(item.get('bidRateCK')),
                        'sell': _to_float(item.get('askRate'))
                    }
    except Exception as e:
        pass
    return None



def clean_rate_val(val):
    if val is None or pd.isna(val):
        return None
    val_str = str(val).strip().replace(",", "")
    if val_str.lower() in ('nan', 'none', 'null', ''):
        return None
    val_str = re.sub(r"[#&\s\+]", "", val_str)
    try:
        num = float(val_str)
        if num < 1000:
            num = num * 1000
        return round(num, 2)
    except:
        return None

def _to_float(val):
    try:
        v = float(str(val).replace(',', ''))
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None

def fetch_vcb_usd_api(date_str_yyyy_mm_dd):
    url = 'https://www.vietcombank.com.vn/api/exchangerates'
    try:
        resp = requests.get(url, params={'date': date_str_yyyy_mm_dd}, headers=HEADERS, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            for item in data.get('Data', []):
                if item.get('currencyCode') == 'USD':
                    return {
                        'buy_cash': _to_float(item.get('cash')),
                        'buy_transfer': _to_float(item.get('transfer')),
                        'sell': _to_float(item.get('sell'))
                    }
    except Exception as e:
        pass
    return None

def fetch_bidv_usd_api(date_str_dd_mm_yyyy):
    url = f"https://bidv.com.vn/ServicesBIDV/ExchangeDetailServlet?date={date_str_dd_mm_yyyy}"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10, verify=False)
        if resp.status_code == 200:
            data = resp.json()
            for item in data.get('data', []):
                if item.get('currency') == 'USD':
                    return {
                        'buy_cash': clean_rate_val(item.get('muaTm')),
                        'buy_transfer': clean_rate_val(item.get('muaCk')),
                        'sell': clean_rate_val(item.get('ban'))
                    }
    except Exception as e:
        pass
    return None

def fetch_acb_usd_api(date_str_yyyy_mm_dd):
    url = f"https://acb.com.vn/api/front/v1/currency?currency=VND&effectiveDateTime={date_str_yyyy_mm_dd}T14:00:00.000"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10, verify=False)
        if resp.status_code == 200:
            data = resp.json()
            buy_cash = None
            buy_transfer = None
            sell = None
            for item in data:
                if item.get('exchangeCurrency') == 'USD':
                    deal_type = item.get('dealType')
                    inst_type = item.get('instrumentType')
                    rate = clean_rate_val(item.get('exchangeRate'))
                    if deal_type == 'BID' and inst_type == 'CASH':
                           buy_cash = rate
                    elif deal_type == 'BID' and inst_type == 'TRANSFER':
                           buy_transfer = rate
                    elif deal_type == 'ASK':
                           sell = rate
            if buy_cash or buy_transfer or sell:
                return {'buy_cash': buy_cash, 'buy_transfer': buy_transfer, 'sell': sell}
    except Exception as e:
        pass
    return None

def main():
    print("🚀 Bắt đầu backfill tỷ giá USD các ngân hàng (API thực tế)...")
    if not os.path.exists(OUTPUT_FILE):
        print(f"❌ File Excel không tồn tại: {OUTPUT_FILE}")
        return

    # Pre-load VietinBank history (2 requests total)
    print("📡 Pre-loading VietinBank history...")
    _load_vietinbank_history()

    # 1. Đọc dữ liệu hiện có
    existing_records = {}
    try:
        wb_check = pd.ExcelFile(OUTPUT_FILE)
        sheet_name = 'Data_TheoDoi_USD'
        for s in ['Data_TheoDoi_USD', 'Data_SoSanh_USD']:
            if s in wb_check.sheet_names:
                sheet_name = s
                break
        df_existing = pd.read_excel(OUTPUT_FILE, sheet_name=sheet_name)
        print(f"📂 Đọc được {len(df_existing)} dòng từ sheet {sheet_name}")
        for idx, row in df_existing.iterrows():
            date_val = row.get('Ngày Cập Nhật')
            bank_val = row.get('Ngân Hàng')
            if pd.isna(date_val) or pd.isna(bank_val):
                continue
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%d/%m/%Y')
            else:
                date_str = str(date_val).strip()
            bank_str = str(bank_val).strip()
            existing_records[(date_str, bank_str)] = {
                'buy_cash': clean_rate_val(row.get('Mua Tiền Mặt')),
                'buy_transfer': clean_rate_val(row.get('Mua Chuyển Khoản')),
                'sell': clean_rate_val(row.get('Bán'))
            }
    except Exception as e:
        print(f"⚠ Không đọc được dữ liệu cũ: {e}")

    # 2. Đọc tỷ giá VCB từ sheet Data
    vcb_rates_from_sheet = {}
    try:
        df_vcb = pd.read_excel(OUTPUT_FILE, sheet_name='Data')
        df_vcb_usd = df_vcb[df_vcb['Mã Ngoại Tệ'] == 'USD']
        for idx, row in df_vcb_usd.iterrows():
            date_val = row.get('Ngày Cập Nhật')
            if pd.isna(date_val):
                continue
            if isinstance(date_val, datetime):
                dt_key_yyyy = date_val.strftime('%Y-%m-%d')
                dt_key_dd = date_val.strftime('%d/%m/%Y')
            else:
                try:
                    dt = pd.to_datetime(date_val)
                    dt_key_yyyy = dt.strftime('%Y-%m-%d')
                    dt_key_dd = dt.strftime('%d/%m/%Y')
                except:
                    dt_key_yyyy = str(date_val)[:10]
                    parts = dt_key_yyyy.split('-')
                    dt_key_dd = f"{parts[2]}/{parts[1]}/{parts[0]}" if len(parts) == 3 else dt_key_yyyy
            rates = {
                'buy_cash': clean_rate_val(row.get('Mua Tiền Mặt')),
                'buy_transfer': clean_rate_val(row.get('Mua Chuyển Khoản')),
                'sell': clean_rate_val(row.get('Bán'))
            }
            vcb_rates_from_sheet[dt_key_yyyy] = rates
            vcb_rates_from_sheet[dt_key_dd] = rates
        print(f"📂 Đọc được VCB USD từ sheet Data: {len(df_vcb_usd)} dòng")
    except Exception as e:
        print(f"⚠ Không đọc được VCB từ sheet Data: {e}")

    # Techcombank today
    print("📡 Đang lấy tỷ giá Techcombank hôm nay...")
    techcom_today = fetch_techcombank_usd_current()
    today_str = END_DATE.strftime('%d/%m/%Y')
    print(f"  Techcombank hôm nay ({today_str}): {techcom_today}")

    # 3. Generate date list
    dates = []
    curr = START_DATE
    while curr <= END_DATE:
        dates.append(curr)
        curr += timedelta(days=1)
    print(f"📅 Xử lý {len(dates)} ngày ({START_DATE.strftime('%d/%m/%Y')} -> {END_DATE.strftime('%d/%m/%Y')})")

    # 4. Thu thập dữ liệu
    new_records = []
    
    def fetch_data_for_date(dt):
        date_dd = dt.strftime('%d/%m/%Y')
        date_yyyy = dt.strftime('%Y-%m-%d')
        is_today = (date_dd == today_str)
        
        # Check existing and valid
        all_exist = True
        for bank in BANKS:
            r = existing_records.get((date_dd, bank))
            if not r or (r.get('buy_transfer') is None and r.get('sell') is None):
                all_exist = False
                break
                
        if all_exist:
            rows = []
            for bank in BANKS:
                r = existing_records[(date_dd, bank)]
                rows.append([date_dd, bank, r['buy_cash'], r['buy_transfer'], r['sell']])
            return rows

        # A. VCB
        vcb_usd = existing_records.get((date_dd, 'Vietcombank'))
        if not vcb_usd or (vcb_usd.get('buy_transfer') is None and vcb_usd.get('sell') is None):
            vcb_usd = vcb_rates_from_sheet.get(date_yyyy) or vcb_rates_from_sheet.get(date_dd)
            if not vcb_usd:
                vcb_usd = fetch_vcb_usd_api(date_yyyy)
                if vcb_usd:
                    vcb_rates_from_sheet[date_yyyy] = vcb_usd
                    vcb_rates_from_sheet[date_dd] = vcb_usd
        if not vcb_usd:
            return []

        # B. BIDV
        bidv_usd = existing_records.get((date_dd, 'BIDV'))
        if not bidv_usd or (bidv_usd.get('buy_transfer') is None and bidv_usd.get('sell') is None):
            bidv_usd = fetch_bidv_usd_api(date_dd)
        if not bidv_usd:
            bidv_usd = {'buy_cash': None, 'buy_transfer': None, 'sell': None}

        # C. ACB
        acb_usd = existing_records.get((date_dd, 'ACB'))
        if not acb_usd or (acb_usd.get('buy_transfer') is None and acb_usd.get('sell') is None):
            acb_usd = fetch_acb_usd_api(date_yyyy)
        if not acb_usd:
            acb_usd = {'buy_cash': None, 'buy_transfer': None, 'sell': None}

        # D. VietinBank (from history cache or current API)
        vietinbank_usd = existing_records.get((date_dd, 'VietinBank'))
        if not vietinbank_usd or (vietinbank_usd.get('buy_transfer') is None and vietinbank_usd.get('sell') is None):
            vietinbank_usd = get_vietinbank_usd_for_date(date_yyyy)
        if not vietinbank_usd and is_today:
            vietinbank_usd = fetch_vietinbank_usd_current(date_yyyy)
        if not vietinbank_usd:
            vietinbank_usd = {'buy_cash': None, 'buy_transfer': None, 'sell': None}

        # E. SeaBank (historical API)
        seabank_usd = existing_records.get((date_dd, 'SeaBank'))
        if not seabank_usd or (seabank_usd.get('buy_transfer') is None and seabank_usd.get('sell') is None):
            seabank_usd = get_seabank_usd_for_date(date_dd)
        if not seabank_usd:
            seabank_usd = {'buy_cash': None, 'buy_transfer': None, 'sell': None}

        # F. Techcombank (historical API)
        techcom_usd = existing_records.get((date_dd, 'Techcombank'))
        if not techcom_usd or (techcom_usd.get('buy_transfer') is None and techcom_usd.get('sell') is None):
            techcom_usd = fetch_techcombank_usd_historical(date_yyyy)
        if not techcom_usd and is_today and techcom_today:
            techcom_usd = techcom_today
        if not techcom_usd:
            techcom_usd = {'buy_cash': None, 'buy_transfer': None, 'sell': None}

        rows = []
        bank_data = {
            'Vietcombank': vcb_usd,
            'VietinBank': vietinbank_usd,
            'BIDV': bidv_usd,
            'Techcombank': techcom_usd,
            'ACB': acb_usd,
            'SeaBank': seabank_usd
        }
        for bank in BANKS:
            d = bank_data[bank]
            rows.append([date_dd, bank, d.get('buy_cash'), d.get('buy_transfer'), d.get('sell')])
        return rows

    print("⏳ Đang thu thập tất cả dữ liệu...")
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(fetch_data_for_date, d): d for d in dates}
        for i, fut in enumerate(concurrent.futures.as_completed(futures)):
            res = fut.result()
            if res:
                new_records.extend(res)
            if (i + 1) % 30 == 0 or (i + 1) == len(dates):
                print(f"  ✅ Hoàn tất {i + 1}/{len(dates)} ngày")

    if not new_records:
        print("❌ Không thu thập được dữ liệu nào!")
        return

    df_new = pd.DataFrame(new_records, columns=['Ngày Cập Nhật', 'Ngân Hàng', 'Mua Tiền Mặt', 'Mua Chuyển Khoản', 'Bán'])
    print(f"💾 Tổng số dòng dữ liệu mới: {len(df_new)}")
    
    # Close Excel if open
    import subprocess, time
    try:
        subprocess.run(['taskkill', '/f', '/im', 'excel.exe'], capture_output=True)
        time.sleep(1.0)
    except:
        pass
        
    wb = load_workbook(OUTPUT_FILE)
    
    for sheet in ['Data_SoSanh_USD', 'SoSanh_USD']:
        if sheet in wb.sheetnames:
            wb.remove(wb[sheet])
            
    if 'Data_TheoDoi_USD' in wb.sheetnames:
        ws_data = wb['Data_TheoDoi_USD']
        ws_data.delete_rows(1, ws_data.max_row + 10)
    else:
        ws_data = wb.create_sheet('Data_TheoDoi_USD')
    ws_data.sheet_state = 'hidden'
    
    col_headers = ['Ngày Cập Nhật', 'Ngân Hàng', 'Mua Tiền Mặt', 'Mua Chuyển Khoản', 'Bán']
    for col_idx, h in enumerate(col_headers, 1):
        ws_data.cell(1, col_idx, h)
        
    def parse_date_key(x):
        try:
            return datetime.strptime(x[0], '%d/%m/%Y')
        except:
            return datetime.min
            
    sorted_rows = sorted(new_records, key=lambda x: (parse_date_key(x), x[1]), reverse=True)
    for r_idx, row_vals in enumerate(sorted_rows, 2):
        for c_idx, val in enumerate(row_vals, 1):
            ws_data.cell(r_idx, c_idx, val)
            
    print("📈 Đang xây dựng Dashboard TheoDoi_USD...")
    sys.path.insert(0, r'D:\Tygia-Tudong\scripts')
    from get_rates import _create_multi_bank_comparison_sheet
    current_date_str = datetime.now().strftime('%d/%m/%Y')
    _create_multi_bank_comparison_sheet(wb, current_date_str)
    
    wb.save(OUTPUT_FILE)
    wb.close()
    print("✅ Đã cập nhật thành công TyGia_Banking.xlsx với dữ liệu API thực tế!")
    
    # Print summary
    print("\n📊 Tóm tắt dữ liệu thu được:")
    summary = {}
    for row in new_records:
        bank = row[1]
        if bank not in summary:
            summary[bank] = {'total': 0, 'with_data': 0}
        summary[bank]['total'] += 1
        if row[3] is not None:  # buy_transfer
            summary[bank]['with_data'] += 1
    for bank, s in summary.items():
        pct = (s['with_data'] / s['total'] * 100) if s['total'] > 0 else 0
        print(f"  {bank}: {s['with_data']}/{s['total']} ngày có dữ liệu ({pct:.0f}%)")

if __name__ == '__main__':
    main()
