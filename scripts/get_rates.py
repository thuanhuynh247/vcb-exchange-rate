import os
os.makedirs(r"D:\Tygia-Tudong\Temp", exist_ok=True)
os.environ["TEMP"] = r"D:\Tygia-Tudong\Temp"
os.environ["TMP"] = r"D:\Tygia-Tudong\Temp"
os.environ["OPENPYXL_LXML"] = "False"
import openpyxl
openpyxl.xml.LXML = False
import requests
import pandas as pd
from datetime import datetime
import xml.etree.ElementTree as ET
import os
import logging
import time
import subprocess
import asyncio
import re
import json
import urllib3
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# Cấu hình
URL = 'https://portal.vietcombank.com.vn/Usercontrols/TVPortal.TyGia/pXML.aspx?b=68'
OUTPUT_FILE = r'D:\Tygia-Tudong\TyGia_Banking.xlsx'
LOG_FILE = r'D:\Tygia-Tudong\vcb_rates.log'
RETRY_COUNT = 3
RETRY_DELAY = 30  # seconds

# Thứ tự hiển thị ngoại tệ chính (giống website VCB)
PRIORITY_CURRENCIES = ['USD', 'EUR', 'GBP', 'JPY', 'AUD', 'CAD', 'CHF', 'SGD',
                       'CNY', 'HKD', 'THB', 'KRW', 'MYR', 'INR', 'KWD',
                       'SAR', 'NOK', 'SEK', 'DKK', 'RUB']

# Ngoại tệ dùng cho biểu đồ Dashboard
CHART_CURRENCIES = ['USD', 'EUR']

# ── Logging setup (chỉ add handler 1 lần) ──
logger = logging.getLogger('VCB_Rates')
logger.setLevel(logging.INFO)
if not logger.handlers:
    _fh = logging.FileHandler(LOG_FILE, encoding='utf-8')
    _fh.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    _ch = logging.StreamHandler()
    _ch.setFormatter(logging.Formatter('%(message)s'))
    logger.addHandler(_fh)
    logger.addHandler(_ch)


def send_notification(title, message):
    """Send Windows toast notification via PowerShell."""
    try:
        ps_script = f'''
[Windows.UI.Notifications.ToastNotificationManager, Windows.UI.Notifications, ContentType = WindowsRuntime] | Out-Null
[Windows.Data.Xml.Dom.XmlDocument, Windows.Data.Xml.Dom, ContentType = WindowsRuntime] | Out-Null
$template = @"
<toast>
    <visual><binding template="ToastText02">
        <text id="1">{title}</text>
        <text id="2">{message}</text>
    </binding></visual>
</toast>
"@
$xml = New-Object Windows.Data.Xml.Dom.XmlDocument
$xml.LoadXml($template)
$toast = [Windows.UI.Notifications.ToastNotification]::new($xml)
[Windows.UI.Notifications.ToastNotificationManager]::CreateToastNotifier("VCB Exchange Rate").Show($toast)
'''
        subprocess.run(['powershell', '-Command', ps_script],
                       capture_output=True, timeout=10)
    except Exception:
        pass  # Notification is best-effort


def get_exchange_rates():
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36'
    }
    try:
        response = requests.get(URL, headers=headers, timeout=10)
        response.raise_for_status()

        root = ET.fromstring(response.content)
        data = []
        update_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        for child in root.findall('Exrate'):
            currency_code = child.get('CurrencyCode', '').strip()
            currency_name = child.get('CurrencyName', '').strip()
            buy = child.get('Buy', '').replace(',', '').strip()
            transfer = child.get('Transfer', '').replace(',', '').strip()
            sell = child.get('Sell', '').replace(',', '').strip()

            data.append({
                'Ngày Cập Nhật': update_time_str,
                'Mã Ngoại Tệ': currency_code,
                'Tên Ngoại Tệ': currency_name,
                'Mua Tiền Mặt': _to_float(buy),
                'Mua Chuyển Khoản': _to_float(transfer),
                'Bán': _to_float(sell)
            })
        return data
    except Exception as e:
        print(f"Lỗi khi lấy dữ liệu: {e}")
        return None


def _to_float(val):
    try:
        v = float(str(val).replace(',', ''))
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None


# =======================================================================
# THU THẬP TỶ GIÁ USD/VND TỪ NHIỀU NGÂN HÀNG (VIETCOMBANK, VIETINBANK, BIDV, TECHCOMBANK, ACB, SEABANK)
# =======================================================================

def clean_rate_val(val):
    if not val:
        return None
    val_str = str(val).strip().replace(",", "")
    val_str = re.sub(r"[#&\s\+]", "", val_str)
    try:
        num = float(val_str)
        if num < 1000:
            num = num * 1000
        return round(num, 2)
    except:
        return None

def fetch_bidv_usd():
    logger.info("Đang lấy tỷ giá BIDV...")
    url = "https://bidv.com.vn/ServicesBIDV/ExchangeDetailServlet"
    try:
        r = requests.post(url, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}, timeout=15, verify=False)
        if r.status_code == 200:
            data = r.json()
            for item in data.get("data", []):
                if item.get("currency") == "USD":
                    return {
                        "bank": "BIDV",
                        "buy_cash": clean_rate_val(item.get("muaTm")),
                        "buy_transfer": clean_rate_val(item.get("muaCk")),
                        "sell": clean_rate_val(item.get("ban"))
                    }
    except Exception as e:
        logger.error(f"Lỗi lấy tỷ giá BIDV: {e}")
    return {"bank": "BIDV", "buy_cash": None, "buy_transfer": None, "sell": None}

def fetch_techcombank_usd():
    logger.info("Đang lấy tỷ giá Techcombank...")
    csrf_url = "https://techcombank.com/libs/granite/csrf/token.json"
    rate_url = "https://techcombank.com/content/techcombank/web/vn/vi/cong-cu-tien-ich/ty-gia/_jcr_content.exchange-rates.integration.json"
    session_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://techcombank.com/cong-cu-tien-ich/ty-gia"
    }
    try:
        session = requests.Session()
        session.get(csrf_url, headers=session_headers, verify=False, timeout=10)  # init CSRF/cookies
        r = session.get(rate_url, headers=session_headers, verify=False, timeout=15)
        if r.status_code == 200:
            data = r.json()
            rates = data.get("exchangeRate", {}).get("data", [])
            for item in rates:
                if item.get("label", "").startswith("USD (50") or item.get("label") == "USD (50,100)":  # prefer large bills
                    return {
                        "bank": "Techcombank",
                        "buy_cash": clean_rate_val(item.get("bidRateTM")),
                        "buy_transfer": clean_rate_val(item.get("bidRateCK")),
                        "sell": clean_rate_val(item.get("askRate"))
                    }
    except Exception as e:
        logger.error(f"Lỗi lấy tỷ giá Techcombank: {e}")
    return {"bank": "Techcombank", "buy_cash": None, "buy_transfer": None, "sell": None}

def fetch_acb_usd():
    logger.info("Đang lấy tỷ giá ACB...")
    try:
        now_str = datetime.now().strftime("%Y-%m-%dT%H:%M:%S.000")
        url = f"https://acb.com.vn/api/front/v1/currency?currency=VND&effectiveDateTime={now_str}"
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}, timeout=15, verify=False)
        if r.status_code == 200:
            data = r.json()
            usd_rates = [x for x in data if x.get("exchangeCurrency") == "USD"]
            buy_cash = None
            buy_transfer = None
            sell = None
            for item in usd_rates:
                deal_type = item.get("dealType")
                inst_type = item.get("instrumentType")
                rate = clean_rate_val(item.get("exchangeRate"))
                if deal_type == "BID" and inst_type == "CASH":
                    buy_cash = rate
                elif deal_type == "BID" and inst_type == "TRANSFER":
                    buy_transfer = rate
                elif deal_type == "ASK":
                    sell = rate
            return {"bank": "ACB", "buy_cash": buy_cash, "buy_transfer": buy_transfer, "sell": sell}
    except Exception as e:
        logger.error(f"Lỗi lấy tỷ giá ACB: {e}")
    return {"bank": "ACB", "buy_cash": None, "buy_transfer": None, "sell": None}

def fetch_vietinbank_usd():
    """Fetch VietinBank current USD rate via Server Action API (no Playwright needed)."""
    logger.info("Đang lấy tỷ giá VietinBank qua Server Action API...")
    url = "https://www.vietinbank.vn/ty-gia-khcn"
    action = "1e43a43a5124d6cc3cb463bc54021b34f39a4065"
    h = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Content-Type": "text/plain;charset=UTF-8",
        "next-action": action,
        "accept": "text/x-component",
        "referer": "https://www.vietinbank.vn/ty-gia-khcn"
    }
    date_str = datetime.now().strftime('%Y-%m-%d')
    try:
        r = requests.post(url, headers=h, data=json.dumps([f"{date_str}T15:45:00", "USD"]),
                         verify=False, timeout=15)
        for line in r.text.strip().split('\n'):
            if line.startswith("1:"):
                data = json.loads(line[2:])
                if data and isinstance(data, list) and len(data) > 0:
                    item = data[0]
                    return {
                        "bank": "VietinBank",
                        "buy_cash": item.get('cash_rate_big'),
                        "buy_transfer": item.get('transfer_rate'),
                        "sell": item.get('sell_rate')
                    }
    except Exception as e:
        logger.error(f"Lỗi VietinBank Server Action: {e}")
        
    # Fallback to history Server Action API
    logger.info("Thử lấy tỷ giá VietinBank qua Server Action Lịch Sử...")
    action_history = "ff24b60505a8da357a655878afe7dd2d1f9f0e52"
    h["next-action"] = action_history
    try:
        # Fetch transfer rate
        r_tf = requests.post(url, headers=h, data=json.dumps([date_str, date_str, "USD", "transfer_rate"]),
                             verify=False, timeout=15)
        transfer_val = None
        for line in r_tf.text.strip().split('\n'):
            if line.startswith("1:"):
                data = json.loads(line[2:])
                if data and isinstance(data, list) and len(data) > 0:
                    transfer_val = data[0].get('close', data[0].get('open'))
                    
        # Fetch sell rate
        r_sell = requests.post(url, headers=h, data=json.dumps([date_str, date_str, "USD", "sell_rate"]),
                               verify=False, timeout=15)
        sell_val = None
        for line in r_sell.text.strip().split('\n'):
            if line.startswith("1:"):
                data = json.loads(line[2:])
                if data and isinstance(data, list) and len(data) > 0:
                    sell_val = data[0].get('close', data[0].get('open'))
                    
        if transfer_val or sell_val:
            return {
                "bank": "VietinBank",
                "buy_cash": transfer_val,
                "buy_transfer": transfer_val,
                "sell": sell_val
            }
    except Exception as e:
        logger.error(f"Lỗi VietinBank Server Action Lịch Sử: {e}")
        
    return {"bank": "VietinBank", "buy_cash": None, "buy_transfer": None, "sell": None}


def fetch_seabank_usd():
    """Fetch SeaBank current USD rate via Server Action API (no Playwright needed)."""
    logger.info("Đang lấy tỷ giá SeaBank qua Server Action API...")
    url = "https://www.seabank.com.vn/cong-cu-tien-ich/ty-gia"
    action = "d65f2411081b93638167328d79ca76cf2bc7ec18"
    h = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Content-Type": "text/plain;charset=UTF-8",
        "next-action": action,
        "accept": "text/x-component",
        "referer": "https://www.seabank.com.vn/cong-cu-tien-ich/ty-gia"
    }
    date_str = datetime.now().strftime('%d/%m/%Y')
    try:
        r = requests.post(url, headers=h, data=json.dumps([date_str]),
                         verify=False, timeout=15)
        for line in r.text.strip().split('\n'):
            if line.startswith("1:"):
                data = json.loads(line[2:])
                if data and isinstance(data, dict) and 'details' in data:
                    usd = next((x for x in data['details'] if x.get('currency') == 'USD'), None)
                    if usd:
                        def to_f(v):
                            try:
                                return float(str(v).replace(',', '')) if v else None
                            except: return None
                        return {
                            "bank": "SeaBank",
                            "buy_cash": to_f(usd.get('buy')),
                            "buy_transfer": to_f(usd.get('transferBuy')),
                            "sell": to_f(usd.get('sell'))
                        }
    except Exception as e:
        logger.error(f"Lỗi SeaBank Server Action: {e}")
    return {"bank": "SeaBank", "buy_cash": None, "buy_transfer": None, "sell": None}

def get_multi_bank_rates(vcb_usd_data=None):
    logger.info("=== BẮT ĐẦU THU THẬP TỶ GIÁ USD CÁC NGÂN HÀNG ===")
    
    vcb_res = {"bank": "Vietcombank", "buy_cash": None, "buy_transfer": None, "sell": None}
    if vcb_usd_data:
        vcb_res["buy_cash"] = vcb_usd_data.get("buy_cash") or vcb_usd_data.get("Mua Tiền Mặt")
        vcb_res["buy_transfer"] = vcb_usd_data.get("buy_transfer") or vcb_usd_data.get("Mua Chuyển Khoản")
        vcb_res["sell"] = vcb_usd_data.get("sell") or vcb_usd_data.get("Bán")
    else:
        try:
            r = requests.get('https://portal.vietcombank.com.vn/Usercontrols/TVPortal.TyGia/pXML.aspx?b=68', timeout=15)
            root = ET.fromstring(r.content)
            for child in root.findall('Exrate'):
                if child.get('CurrencyCode', '').strip() == 'USD':
                    vcb_res["buy_cash"] = clean_rate_val(child.get('Buy'))
                    vcb_res["buy_transfer"] = clean_rate_val(child.get('Transfer'))
                    vcb_res["sell"] = clean_rate_val(child.get('Sell'))
                    break
        except Exception as e:
            logger.error(f"Lỗi lấy VCB USD: {e}")
            
    bidv = fetch_bidv_usd()
    tcb = fetch_techcombank_usd()
    acb = fetch_acb_usd()
    vtb = fetch_vietinbank_usd()
    seab = fetch_seabank_usd()
        
    res_list = [vcb_res, vtb, bidv, tcb, acb, seab]
    logger.info(f"Kết quả thu thập: {res_list}")
    return res_list

def _save_multi_bank_data(wb, multi_bank_data, current_date_str):
    if 'Data_TheoDoi_USD' not in wb.sheetnames:
        ws_data = wb.create_sheet('Data_TheoDoi_USD')
    else:
        ws_data = wb['Data_TheoDoi_USD']
        
    ws_data.sheet_state = 'hidden'
    
    existing_rows = []
    if ws_data.max_row > 1:
        for r in range(2, ws_data.max_row + 1):
            date_val = ws_data.cell(r, 1).value
            bank_val = ws_data.cell(r, 2).value
            buy_cash = ws_data.cell(r, 3).value
            buy_transfer = ws_data.cell(r, 4).value
            sell = ws_data.cell(r, 5).value
            if date_val and bank_val:
                existing_rows.append([str(date_val).strip(), str(bank_val).strip(), buy_cash, buy_transfer, sell])
                
    if not existing_rows and 'Data' in wb.sheetnames:
        logger.info("Bắt đầu di trú lịch sử tỷ giá USD của Vietcombank sang Data_TheoDoi_USD...")
        ws_vcb = wb['Data']
        vcb_rows = []
        for r in range(2, ws_vcb.max_row + 1):
            code = ws_vcb.cell(r, 2).value
            if code == 'USD':
                date_val = ws_vcb.cell(r, 1).value
                if not date_val:
                    continue
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime('%d/%m/%Y')
                else:
                    try:
                        dt = pd.to_datetime(date_val)
                        date_str = dt.strftime('%d/%m/%Y')
                    except:
                        date_str = str(date_val)[:10]
                
                buy_cash = ws_vcb.cell(r, 4).value
                buy_transfer = ws_vcb.cell(r, 5).value
                sell = ws_vcb.cell(r, 6).value
                vcb_rows.append([date_str, 'Vietcombank', buy_cash, buy_transfer, sell])
        existing_rows.extend(vcb_rows)
        logger.info(f"Đã di trú {len(vcb_rows)} dòng lịch sử VCB.")
        
    today_str = datetime.now().strftime('%d/%m/%Y')
    for item in multi_bank_data:
        if item['bank'] == 'Vietcombank':
            row_date = current_date_str
        else:
            row_date = today_str
            
        existing_rows.append([row_date, item['bank'], item['buy_cash'], item['buy_transfer'], item['sell']])
        
    unique_map = {}
    for r in existing_rows:
        key = (r[0], r[1])
        unique_map[key] = r
        
    def parse_date_key(x):
        try:
            return datetime.strptime(x[0], '%d/%m/%Y')
        except:
            return datetime.min
            
    df_temp = pd.DataFrame(list(unique_map.values()), columns=['Ngày Cập Nhật', 'Ngân Hàng', 'Mua Tiền Mặt', 'Mua Chuyển Khoản', 'Bán'])
    df_temp['_parsed_date'] = df_temp['Ngày Cập Nhật'].apply(parse_date_key)
    
    # Sort chronologically to forward fill
    df_temp = df_temp.sort_values(by=['Ngân Hàng', '_parsed_date'])
    df_temp['Mua Tiền Mặt'] = df_temp.groupby('Ngân Hàng')['Mua Tiền Mặt'].ffill().bfill()
    df_temp['Mua Chuyển Khoản'] = df_temp.groupby('Ngân Hàng')['Mua Chuyển Khoản'].ffill().bfill()
    df_temp['Bán'] = df_temp.groupby('Ngân Hàng')['Bán'].ffill().bfill()
    
    # Sort back to descending order
    df_temp = df_temp.sort_values(by=['_parsed_date', 'Ngân Hàng'], ascending=[False, True])
    df_temp.drop(columns=['_parsed_date'], inplace=True)
    
    ws_data.delete_rows(1, ws_data.max_row + 10)
    
    headers = ['Ngày Cập Nhật', 'Ngân Hàng', 'Mua Tiền Mặt', 'Mua Chuyển Khoản', 'Bán']
    for col_idx, h in enumerate(headers, 1):
        ws_data.cell(1, col_idx, h)
        
    for r_idx, row_vals in enumerate(df_temp.values, 2):
        for c_idx, val in enumerate(row_vals, 1):
            if pd.isna(val):
                ws_data.cell(r_idx, c_idx, None)
            else:
                ws_data.cell(r_idx, c_idx, val)

def _create_multi_bank_comparison_sheet(wb, current_date_str):
    if 'TheoDoi_USD' in wb.sheetnames:
        ws = wb['TheoDoi_USD']
        ws.delete_rows(1, ws.max_row + 100)
    else:
        ws = wb.create_sheet('TheoDoi_USD', 1)
        
    ws.sheet_view.showGridLines = False
    
    THEME_GREEN = '00703C'
    THEME_GREEN_LIGHT = 'E8F5E9'
    THEME_WHITE = 'FFFFFF'
    THEME_DARK = '333333'
    THEME_GRAY = '666666'
    THEME_BLUE = '1565C0'
    THEME_BLUE_LIGHT = 'E3F2FD'
    
    font_title = Font(bold=True, size=18, color=THEME_DARK, name='Arial')
    font_header = Font(bold=True, color=THEME_WHITE, size=11, name='Arial')
    font_body = Font(size=11, name='Arial')
    font_bold = Font(bold=True, size=11, name='Arial')
    
    fill_header_green = PatternFill(start_color=THEME_GREEN, end_color=THEME_GREEN, fill_type='solid')
    fill_header_blue = PatternFill(start_color=THEME_BLUE, end_color=THEME_BLUE, fill_type='solid')
    fill_even_row = PatternFill(start_color=THEME_GREEN_LIGHT, end_color=THEME_GREEN_LIGHT, fill_type='solid')
    fill_compare_blue = PatternFill(start_color=THEME_BLUE_LIGHT, end_color=THEME_BLUE_LIGHT, fill_type='solid')
    
    border_thin = Border(
        left=Side(style='thin', color='C8E6C9'),
        right=Side(style='thin', color='C8E6C9'),
        top=Side(style='thin', color='C8E6C9'),
        bottom=Side(style='thin', color='C8E6C9'),
    )
    border_picker = Border(
        left=Side(style='medium', color=THEME_GREEN),
        right=Side(style='medium', color=THEME_GREEN),
        top=Side(style='medium', color=THEME_GREEN),
        bottom=Side(style='medium', color=THEME_GREEN),
    )
    border_picker_blue = Border(
        left=Side(style='medium', color=THEME_BLUE),
        right=Side(style='medium', color=THEME_BLUE),
        top=Side(style='medium', color=THEME_BLUE),
        bottom=Side(style='medium', color=THEME_BLUE),
    )
    
    col_widths = {
        'A': 3, 'B': 22, 'C': 18, 'D': 20, 'E': 18, 'F': 22, 'G': 18, 'H': 18, 'I': 16, 'J': 3
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
        
    ws.merge_cells('B2:I2')
    ws['B2'].value = 'BẢNG SO SÁNH TỶ GIÁ USD/VND GIỮA CÁC NGÂN HÀNG'
    ws['B2'].font = font_title
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 40
    
    ws['B3'].value = 'Ngày xem'
    ws['B3'].font = Font(size=10, color=THEME_GRAY, name='Arial')
    ws['E3'].value = 'So sánh với ngày'
    ws['E3'].font = Font(size=10, color=THEME_GRAY, name='Arial')
    
    ws['B4'].value = current_date_str
    ws['B4'].font = Font(size=13, color=THEME_DARK, name='Arial', bold=True)
    ws['B4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B4'].border = border_picker
    ws['B4'].fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
    
    ws['C4'].value = '📅 Chọn ngày'
    ws['C4'].font = Font(size=10, color=THEME_GRAY, italic=True, name='Arial')
    ws['C4'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws_data = wb['Data_TheoDoi_USD']
    unique_dates = []
    if ws_data.max_row > 1:
        for r in range(2, ws_data.max_row + 1):
            d = ws_data.cell(r, 1).value
            if d:
                unique_dates.append(str(d))
    unique_dates = sorted(list(set(unique_dates)), key=lambda x: datetime.strptime(x, '%d/%m/%Y'), reverse=True)
    
    compare_date_str = current_date_str
    if len(unique_dates) > 1:
        compare_date_str = unique_dates[1] if unique_dates[0] == current_date_str else unique_dates[0]
        
    ws['E4'].value = compare_date_str
    ws['E4'].font = Font(size=13, color=THEME_BLUE, name='Arial', bold=True)
    ws['E4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E4'].border = border_picker_blue
    ws['E4'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    
    ws['F4'].value = '📅 So sánh'
    ws['F4'].font = Font(size=10, color=THEME_GRAY, italic=True, name='Arial')
    ws['F4'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[4].height = 32
    
    if unique_dates:
        for i, d_str in enumerate(unique_dates, 1):
            ws.cell(i, 27, d_str)
        ws.column_dimensions['AA'].hidden = True
        
        date_range_ref = f"=$AA$1:$AA${len(unique_dates)}"
        dv = DataValidation(type="list", formula1=date_range_ref, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws['B4'])
        dv.add(ws['E4'])
        
    table_headers = [
        'Ngân hàng', 'Mua tiền mặt', 'Mua chuyển khoản', 'Bán', 
        'Chênh lệch Mua-Bán', 'Bán (SS)', 'Chênh lệch', '% So sánh'
    ]
    table_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    for col_letter, text in zip(table_cols, table_headers):
        cell = ws[f'{col_letter}7']
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header_blue if col_letter in ('G', 'H', 'I') else fill_header_green
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
    ws.row_dimensions[7].height = 35
    
    banks = ["Vietcombank", "VietinBank", "BIDV", "Techcombank", "ACB", "SeaBank"]
    
    for i, bank in enumerate(banks):
        row_idx = 8 + i
        ws.row_dimensions[row_idx].height = 28
        
        ws[f'B{row_idx}'].value = bank
        ws[f'B{row_idx}'].font = font_bold
        ws[f'B{row_idx}'].border = border_thin
        ws[f'B{row_idx}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws[f'C{row_idx}'].value = f'=IF(SUMIFS(Data_TheoDoi_USD!$C$2:$C$10000, Data_TheoDoi_USD!$A$2:$A$10000, $B$4, Data_TheoDoi_USD!$B$2:$B$10000, $B{row_idx})=0, "-", SUMIFS(Data_TheoDoi_USD!$C$2:$C$10000, Data_TheoDoi_USD!$A$2:$A$10000, $B$4, Data_TheoDoi_USD!$B$2:$B$10000, $B{row_idx}))'
        ws[f'C{row_idx}'].font = font_body
        ws[f'C{row_idx}'].border = border_thin
        ws[f'C{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'C{row_idx}'].number_format = '#,##0.00'
        
        ws[f'D{row_idx}'].value = f'=IF(SUMIFS(Data_TheoDoi_USD!$D$2:$D$10000, Data_TheoDoi_USD!$A$2:$A$10000, $B$4, Data_TheoDoi_USD!$B$2:$B$10000, $B{row_idx})=0, "-", SUMIFS(Data_TheoDoi_USD!$D$2:$D$10000, Data_TheoDoi_USD!$A$2:$A$10000, $B$4, Data_TheoDoi_USD!$B$2:$B$10000, $B{row_idx}))'
        ws[f'D{row_idx}'].font = font_body
        ws[f'D{row_idx}'].border = border_thin
        ws[f'D{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'D{row_idx}'].number_format = '#,##0.00'
        
        ws[f'E{row_idx}'].value = f'=IF(SUMIFS(Data_TheoDoi_USD!$E$2:$E$10000, Data_TheoDoi_USD!$A$2:$A$10000, $B$4, Data_TheoDoi_USD!$B$2:$B$10000, $B{row_idx})=0, "-", SUMIFS(Data_TheoDoi_USD!$E$2:$E$10000, Data_TheoDoi_USD!$A$2:$A$10000, $B$4, Data_TheoDoi_USD!$B$2:$B$10000, $B{row_idx}))'
        ws[f'E{row_idx}'].font = Font(bold=True, color='C0392B', name='Arial')
        ws[f'E{row_idx}'].border = border_thin
        ws[f'E{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'E{row_idx}'].number_format = '#,##0.00'
        
        ws[f'F{row_idx}'].value = f'=IF(AND(ISNUMBER(E{row_idx}), ISNUMBER(D{row_idx})), E{row_idx}-D{row_idx}, "-")'
        ws[f'F{row_idx}'].font = font_body
        ws[f'F{row_idx}'].border = border_thin
        ws[f'F{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'F{row_idx}'].number_format = '#,##0.00'
        
        ws[f'G{row_idx}'].value = f'=IF(SUMIFS(Data_TheoDoi_USD!$E$2:$E$10000, Data_TheoDoi_USD!$A$2:$A$10000, $E$4, Data_TheoDoi_USD!$B$2:$B$10000, $B{row_idx})=0, "-", SUMIFS(Data_TheoDoi_USD!$E$2:$E$10000, Data_TheoDoi_USD!$A$2:$A$10000, $E$4, Data_TheoDoi_USD!$B$2:$B$10000, $B{row_idx}))'
        ws[f'G{row_idx}'].font = font_body
        ws[f'G{row_idx}'].border = border_thin
        ws[f'G{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'G{row_idx}'].number_format = '#,##0.00'
        ws[f'G{row_idx}'].fill = fill_compare_blue
        
        ws[f'H{row_idx}'].value = f'=IF(AND(ISNUMBER(E{row_idx}), ISNUMBER(G{row_idx})), E{row_idx}-G{row_idx}, "-")'
        ws[f'H{row_idx}'].font = font_bold
        ws[f'H{row_idx}'].border = border_thin
        ws[f'H{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'H{row_idx}'].number_format = '+#,##0.00;-#,##0.00;0.00'
        ws[f'H{row_idx}'].fill = fill_compare_blue
        
        ws[f'I{row_idx}'].value = f'=IF(AND(ISNUMBER(E{row_idx}), ISNUMBER(G{row_idx}), G{row_idx}<>0), (E{row_idx}-G{row_idx})/G{row_idx}*100, "-")'
        ws[f'I{row_idx}'].font = font_bold
        ws[f'I{row_idx}'].border = border_thin
        ws[f'I{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'I{row_idx}'].number_format = '+0.00"%";-0.00"%";0.00"%"'
        ws[f'I{row_idx}'].fill = fill_compare_blue
        
        if i % 2 == 1:
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row_idx}'].fill = fill_even_row
            for col in ['G', 'H', 'I']:
                ws[f'{col}{row_idx}'].fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
                
    green_font = Font(bold=True, size=11, color='27AE60', name='Arial')
    red_font = Font(bold=True, size=11, color='C0392B', name='Arial')
    best_buy_fill = PatternFill(start_color=THEME_GREEN_LIGHT, end_color=THEME_GREEN_LIGHT, fill_type='solid')
    best_buy_font = Font(bold=True, color='2E7D32', name='Arial', size=11)
    best_sell_fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
    best_sell_font = Font(bold=True, color='827717', name='Arial', size=11)
    
    ws.conditional_formatting.add('C8:C13', FormulaRule(formula=['C8=MAX($C$8:$C$13)'], fill=best_buy_fill, font=best_buy_font))
    ws.conditional_formatting.add('D8:D13', FormulaRule(formula=['D8=MAX($D$8:$D$13)'], fill=best_buy_fill, font=best_buy_font))
    ws.conditional_formatting.add('E8:E13', FormulaRule(formula=['E8=MIN($E$8:$E$13)'], fill=best_sell_fill, font=best_sell_font))
    
    ws.conditional_formatting.add('H8:H13', CellIsRule(operator='greaterThan', formula=['0'], font=red_font))
    ws.conditional_formatting.add('H8:H13', CellIsRule(operator='lessThan', formula=['0'], font=green_font))
    ws.conditional_formatting.add('I8:I13', CellIsRule(operator='greaterThan', formula=['0'], font=red_font))
    ws.conditional_formatting.add('I8:I13', CellIsRule(operator='lessThan', formula=['0'], font=green_font))
    
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "So Sánh Tỷ Giá USD/VND Giữa Các Ngân Hàng"
    chart.y_axis.title = "VND"
    chart.x_axis.title = "Ngân Hàng"
    chart.width = 20
    chart.height = 12
    
    data_ref = Reference(ws, min_col=3, min_row=7, max_col=5, max_row=13)
    cats_ref = Reference(ws, min_col=2, min_row=8, max_row=13)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    
    colors = ['4CAF50', '2196F3', 'F44336']
    for idx, col in enumerate(colors):
        if idx < len(chart.series):
            chart.series[idx].graphicalProperties.solidFill = col
            
    chart.y_axis.scaling.min = 23000
    chart.y_axis.scaling.max = 27000
    
    ws.add_chart(chart, "B16")
    
    source_row = 32
    ws.merge_cells(f'B{source_row}:I{source_row}')
    ws[f'B{source_row}'].value = "Nguồn: Vietcombank, Vietinbank, BIDV, Techcombank, ACB, SeaBank"
    ws[f'B{source_row}'].font = Font(italic=True, size=9, color='999999', name='Arial')
    ws[f'B{source_row}'].alignment = Alignment(horizontal='center')


def backup_excel(filename):
    if os.path.exists(filename) and os.path.getsize(filename) > 0:
        try:
            backup_dir = os.path.join(os.path.dirname(filename), 'Backup')
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_file = os.path.join(backup_dir, f"TyGia_Banking_Backup_{timestamp}.xlsx")
            import shutil
            shutil.copy2(filename, backup_file)
            logger.info(f"📦 Đã tạo bản sao lưu: {backup_file}")
        except Exception as e:
            logger.warning(f"⚠ Lỗi sao lưu: {e}")


def close_excel_if_file_open(filename):
    """Kiểm tra xem file Excel có đang bị khóa (đang mở) hay không. Nếu có, đóng Excel để giải phóng khóa."""
    if not os.path.exists(filename):
        return
    
    file_locked = False
    try:
        os.rename(filename, filename)
    except OSError:
        file_locked = True
        
    if file_locked:
        logger.info(f"Phát hiện file {os.path.basename(filename)} đang mở. Tiến hành đóng Excel...")
        try:
            subprocess.run(['taskkill', '/f', '/im', 'excel.exe'], capture_output=True)
            time.sleep(1.5)  # Đợi 1.5 giây để Excel giải phóng tài nguyên hoàn toàn
            logger.info("Đã tắt Excel thành công.")
        except Exception as e:
            logger.warning(f"Lỗi khi tắt Excel: {e}")


def _restore_latest_backup(filename):
    """Tìm bản sao lưu gần nhất có dung lượng > 0, kiểm tra tính hợp lệ bằng openpyxl và khôi phục nó."""
    import shutil
    try:
        backup_dir = os.path.join(os.path.dirname(filename), 'Backup')
        if not os.path.exists(backup_dir):
            return False
            
        backups = []
        for f in os.listdir(backup_dir):
            if f.startswith('TyGia_Banking_Backup_') and f.endswith('.xlsx'):
                file_path = os.path.join(backup_dir, f)
                size = os.path.getsize(file_path)
                if size > 0:
                    backups.append((file_path, os.path.getmtime(file_path)))
                    
        if not backups:
            return False
            
        backups.sort(key=lambda x: x[1], reverse=True)
        
        for backup_path, _ in backups:
            try:
                # Kiểm tra tính toàn vẹn của backup
                wb_check = load_workbook(backup_path, read_only=True)
                wb_check.close()
                # Nếu hợp lệ, tiến hành khôi phục
                logger.info(f"Đang khôi phục từ bản sao lưu hợp lệ: {os.path.basename(backup_path)} ({os.path.getsize(backup_path)} bytes)")
                shutil.copy2(backup_path, filename)
                return True
            except Exception as e_check:
                logger.warning(f"Bản sao lưu {os.path.basename(backup_path)} bị lỗi cấu trúc ({e_check}). Thử bản tiếp theo...")
        
        return False
    except Exception as e:
        logger.error(f"Lỗi khi khôi phục bản sao lưu: {e}")
        return False


def _clean_directory(dir_path):
    """Xóa tất cả các file và thư mục con trong thư mục chỉ định, bỏ qua các file bị khóa."""
    import shutil
    try:
        items = os.listdir(dir_path)
    except Exception as e:
        logger.warning(f"Không thể liệt kê thư mục {dir_path}: {e}")
        return

    for item in items:
        item_path = os.path.join(dir_path, item)
        try:
            if os.path.isdir(item_path):
                shutil.rmtree(item_path)
            else:
                os.remove(item_path)
        except Exception:
            pass


def _clean_old_backups(keep_count=30):
    """Chỉ giữ lại một số lượng bản sao lưu nhất định để tránh tốn dung lượng ổ đĩa."""
    try:
        backup_dir = os.path.join(os.path.dirname(OUTPUT_FILE), 'Backup')
        if not os.path.exists(backup_dir):
            return
        
        backups = []
        for f in os.listdir(backup_dir):
            if f.startswith('TyGia_Banking_Backup_') and f.endswith('.xlsx'):
                file_path = os.path.join(backup_dir, f)
                backups.append((file_path, os.path.getmtime(file_path)))
                
        backups.sort(key=lambda x: x[1], reverse=True)
        
        if len(backups) > keep_count:
            logger.info(f"Dọn dẹp bản sao lưu cũ: giữ lại {keep_count} bản mới nhất, xóa {len(backups) - keep_count} bản cũ.")
            for file_path, _ in backups[keep_count:]:
                try:
                    os.remove(file_path)
                except Exception as e:
                    logger.warning(f"Không thể xóa bản sao lưu cũ {os.path.basename(file_path)}: {e}")
    except Exception as e:
        logger.warning(f"Lỗi khi dọn dẹp các bản sao lưu cũ: {e}")


def check_disk_space_and_clean():
    """Kiểm tra dung lượng ổ D, nếu thấp thì dọn dẹp thư mục Temp và các bản sao lưu cũ để đảm bảo chạy tự động."""
    import shutil
    try:
        total, used, free = shutil.disk_usage('D:')
        free_mb = free / (1024 * 1024)
        logger.info(f"Dung lượng trống ổ D: {free_mb:.2f} MB")
        
        if free_mb < 500:
            logger.warning(f"CẢNH BÁO: Dung lượng trống ổ D quá thấp ({free_mb:.2f} MB)! Tiến hành dọn dẹp thư mục Temp...")
            
            temp_path = os.environ.get('TEMP')
            if temp_path and os.path.exists(temp_path):
                _clean_directory(temp_path)
                
            win_temp = r'C:\Windows\Temp'
            if os.path.exists(win_temp):
                _clean_directory(win_temp)
                
            _clean_old_backups(keep_count=20)
            
            _, _, free_new = shutil.disk_usage('D:')
            logger.info(f"Dung lượng trống sau dọn dẹp: {free_new / (1024 * 1024):.2f} MB")
        else:
            _clean_old_backups(keep_count=50)
    except Exception as e:
        logger.error(f"Lỗi khi kiểm tra/dọn dẹp dung lượng ổ đĩa: {e}")


def save_to_excel(data, filename):
    if not data:
        return

    # Tắt file Excel nếu đang mở để tránh bị lock khi ghi
    close_excel_if_file_open(filename)
    
    # Kiểm tra dung lượng ổ đĩa và dọn dẹp trước khi ghi
    check_disk_space_and_clean()

    df_new = pd.DataFrame(data)
    df_combined = df_new
    has_existing = False

    if os.path.exists(filename):
        # Kiểm tra file bị 0 bytes hoặc bị lỗi cấu trúc zip (openpyxl không load được)
        file_valid = True
        if os.path.getsize(filename) == 0:
            file_valid = False
        else:
            try:
                wb_check = load_workbook(filename, read_only=True)
                wb_check.close()
            except Exception as e_check:
                logger.warning(f"Phát hiện file chính bị lỗi cấu trúc ({e_check}). Đang tiến hành khôi phục từ bản sao lưu...")
                file_valid = False
        
        if not file_valid:
            _restore_latest_backup(filename)

        try:
            df_existing = pd.read_excel(filename, sheet_name='Data')
            has_existing = True
        except Exception as e:
            logger.warning(f"Lỗi đọc Data sheet ({e}). Thử khôi phục từ bản sao lưu gần nhất...")
            if _restore_latest_backup(filename):
                try:
                    df_existing = pd.read_excel(filename, sheet_name='Data')
                    has_existing = True
                except Exception as e_inner:
                    logger.error(f"Không thể đọc dữ liệu từ bản sao lưu ({e_inner}). Dùng dữ liệu mới.")
            else:
                logger.error("Không có bản sao lưu hợp lệ để khôi phục. Dùng dữ liệu mới.")

        if has_existing:
            try:
                # Normalize dates and deduplicate
                df_existing['_date_key'] = pd.to_datetime(df_existing['Ngày Cập Nhật'], format='mixed').dt.strftime('%Y-%m-%d')
                df_new['_date_key'] = pd.to_datetime(df_new['Ngày Cập Nhật'], format='mixed').dt.strftime('%Y-%m-%d')

                df_combined = pd.concat([df_new, df_existing], ignore_index=True)
                df_combined = df_combined.drop_duplicates(subset=['_date_key', 'Mã Ngoại Tệ'], keep='first')
                df_combined = df_combined.sort_values(['_date_key', 'Mã Ngoại Tệ'], ascending=[False, True]).reset_index(drop=True)
                df_combined.drop(columns=['_date_key'], inplace=True)
                logger.info(f"Đã đọc {len(df_existing)} dòng cũ, tổng hợp thành {len(df_combined)} dòng.")
            except Exception as e:
                logger.error(f"Lỗi khi gộp dữ liệu cũ và mới ({e}). Dùng dữ liệu mới.")
                df_combined = df_new

    # Đảm bảo có cột Tên Ngoại Tệ
    if 'Tên Ngoại Tệ' not in df_combined.columns:
        df_combined.insert(2, 'Tên Ngoại Tệ', '')

    # Tính bình quân của Mua Chuyển Khoản và Bán
    df_combined['Bình Quân CK & Bán'] = df_combined[['Mua Chuyển Khoản', 'Bán']].mean(axis=1).round(2)

    # Sử dụng file tạm để ghi (ghi thành công toàn bộ mới rename đè lên file chính)
    temp_filename = filename.replace('.xlsx', '_temp.xlsx')
    if os.path.exists(temp_filename):
        try:
            os.remove(temp_filename)
        except Exception:
            pass

    try:
        # Bước 1: Ghi sheet Data vào file tạm
        logger.info("Bước 1/3: Ghi sheet Data...")
        import shutil
        if os.path.exists(filename) and os.path.getsize(filename) > 0:
            shutil.copy2(filename, temp_filename)
            logger.info("Preserved existing sheets by copying file to temp_filename.")
            # Overwrite only the Data sheet
            wb = load_workbook(temp_filename)
            if 'Data' in wb.sheetnames:
                ws_data = wb['Data']
                ws_data.delete_rows(1, ws_data.max_row + 10)
            else:
                ws_data = wb.create_sheet('Data')
            
            # Write df_combined to ws_data
            # Headers
            for col_idx, col_name in enumerate(df_combined.columns, 1):
                ws_data.cell(row=1, column=col_idx, value=col_name)
            # Data rows
            for row_idx, row_vals in enumerate(df_combined.values, 2):
                for col_idx, val in enumerate(row_vals, 1):
                    if pd.isna(val):
                        ws_data.cell(row=row_idx, column=col_idx, value=None)
                    else:
                        ws_data.cell(row=row_idx, column=col_idx, value=val)
            wb.save(temp_filename)
            wb.close()
        else:
            with pd.ExcelWriter(temp_filename, engine='openpyxl') as writer:
                df_combined.to_excel(writer, sheet_name='Data', index=False)

        # Bước 2: Format Data sheet
        logger.info("Bước 2/3: Format Data sheet...")
        wb = load_workbook(temp_filename)
        _format_data_sheet(wb['Data'])
        wb.save(temp_filename)

        # Bước 3: Tạo Dashboard trên file tạm
        logger.info("Bước 3: Tạo Dashboard...")
        wb = load_workbook(temp_filename)
        _create_dashboard_with_date_picker(wb, df_combined)
        wb.save(temp_filename)
        wb.close()

        # Bước 4: Thu thập và so sánh tỷ giá USD các ngân hàng
        logger.info("Bước 4: Thu thập và so sánh tỷ giá USD các ngân hàng...")
        try:
            vcb_usd_data = None
            if data:
                for row in data:
                    if row.get('Mã Ngoại Tệ') == 'USD':
                        vcb_usd_data = {
                            "buy_cash": row.get('Mua Tiền Mặt'),
                            "buy_transfer": row.get('Mua Chuyển Khoản'),
                            "sell": row.get('Bán')
                        }
                        break
            
            multi_bank_data = get_multi_bank_rates(vcb_usd_data)
            
            current_date_str = datetime.now().strftime('%d/%m/%Y')
            if vcb_usd_data and len(data) > 0 and 'Ngày Cập Nhật' in data[0]:
                try:
                    dt = pd.to_datetime(data[0]['Ngày Cập Nhật'])
                    current_date_str = dt.strftime('%d/%m/%Y')
                except Exception as ex_dt:
                    logger.warning(f"Không thể parse Ngày Cập Nhật: {ex_dt}")
            
            wb = load_workbook(temp_filename)
            _save_multi_bank_data(wb, multi_bank_data, current_date_str)
            _create_multi_bank_comparison_sheet(wb, current_date_str)
            wb.save(temp_filename)
            wb.close()
            logger.info("Đã tạo và cập nhật sheet so sánh tỷ giá USD.")
        except Exception as e_multi:
            logger.error(f"Lỗi khi thu thập/so sánh tỷ giá các ngân hàng: {e_multi}")
            import traceback
            logger.error(traceback.format_exc())

        # Tạo backup cho file chính hiện tại TRƯỚC khi đè file tạm lên
        # (Chỉ tạo backup nếu file chính tồn tại và dung lượng > 0 để tránh lưu backup lỗi)
        if os.path.exists(filename) and os.path.getsize(filename) > 0:
            backup_excel(filename)

        # Đè file tạm lên file chính một cách an toàn bằng os.replace (tránh mất file gốc nếu rename lỗi)
        if os.path.exists(filename):
            try:
                os.replace(temp_filename, filename)
            except OSError:
                # Nếu bị lock, cố gắng tắt Excel và thử lại
                close_excel_if_file_open(filename)
                os.replace(temp_filename, filename)
        else:
            os.replace(temp_filename, filename)

        logger.info(f"✅ Đã cập nhật xong Dashboard: {filename}")

    except Exception as e:
        logger.error(f"❌ Lỗi ghi file Excel: {e}")
        import traceback
        logger.error(traceback.format_exc())
        if os.path.exists(temp_filename):
            try:
                os.remove(temp_filename)
            except Exception:
                pass


# ═══════════════════════════════════════════════════════════════════════
# FORMAT DATA SHEET - Dễ tìm kiếm với AutoFilter
# ═══════════════════════════════════════════════════════════════════════
def _format_data_sheet(ws):
    # VCB green palette
    header_fill = PatternFill(start_color='00703C', end_color='00703C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC'),
    )
    even_fill = PatternFill(start_color='F0F8F0', end_color='F0F8F0', fill_type='solid')

    # Format header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[1].height = 32

    # Format data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.font = Font(size=11, name='Arial')
            if cell.column >= 4:  # numeric columns (Mua TM, Mua CK, Bán, Bình Quân CK & Bán)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Alternating row colors
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = even_fill

    # Auto-fit columns
    col_widths = {'A': 22, 'B': 12, 'C': 24, 'D': 18, 'E': 20, 'F': 14, 'G': 20}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header & Add AutoFilter
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════════════
# DASHBOARD VỚI DATE PICKER - Chọn ngày, dữ liệu tự đổi
# ═══════════════════════════════════════════════════════════════════════
def _create_dashboard_with_date_picker(wb, df):
    """Create VCB-style Dashboard with date selection dropdown.
    
    Architecture:
    - Sheet '_Dates': danh sách ngày unique (cho Data Validation)
    - Sheet '_RateData': dữ liệu tỷ giá theo ngày × ngoại tệ (cho INDEX/MATCH)
    - Sheet 'Dashboard': date picker B4 + bảng tỷ giá dùng formula trỏ đến _RateData
    """
    
    # ── Chuẩn bị dữ liệu ──
    df_work = df.copy()
    df_work['_date_key'] = pd.to_datetime(df_work['Ngày Cập Nhật'], format='mixed').dt.strftime('%Y-%m-%d')
    
    # Danh sách ngày unique, sắp xếp giảm dần (mới nhất trước)
    unique_dates = sorted(df_work['_date_key'].unique(), reverse=True)
    # Format dd/mm/yyyy cho display
    display_dates = []
    for d in unique_dates:
        try:
            dt = datetime.strptime(d, '%Y-%m-%d')
            display_dates.append(dt.strftime('%d/%m/%Y'))
        except:
            display_dates.append(d)
    
    latest_date = unique_dates[0] if unique_dates else datetime.now().strftime('%Y-%m-%d')
    latest_display = display_dates[0] if display_dates else datetime.now().strftime('%d/%m/%Y')
    
    # Sắp xếp ngoại tệ theo priority
    def currency_sort_key(code):
        try:
            return PRIORITY_CURRENCIES.index(code)
        except ValueError:
            return 999
    
    all_currencies = sorted(df_work['Mã Ngoại Tệ'].unique(), key=currency_sort_key)
    
    # (The lists will be written directly into the 'Dashboard' sheet hidden columns)
    
    # ═══════════════════════════════════════════════════════
    # SHEET _RateData: Ma trận ngày × ngoại tệ
    # Layout: Col A = ngày (dd/mm/yyyy)
    #         Col B onwards = Mã Ngoại Tệ headers
    #         Mỗi ngoại tệ chiếm 4 cột: Tên, Mua TM, Mua CK, Bán
    # ═══════════════════════════════════════════════════════
    if '_RateData' in wb.sheetnames:
        del wb['_RateData']
    ws_rate = wb.create_sheet('_RateData')
    ws_rate.sheet_state = 'hidden'
    
    # Header row: Ngày | DateSerial | USD_Name | USD_MuaTM | USD_MuaCK | USD_Ban | EUR_Name | ...
    ws_rate.cell(1, 1, 'Ngày')
    ws_rate.cell(1, 2, 'DateSerial')
    col_idx = 3
    currency_col_map = {}  # {currency_code: start_col_index}
    for curr in all_currencies:
        currency_col_map[curr] = col_idx
        ws_rate.cell(1, col_idx, f'{curr}_Name')
        ws_rate.cell(1, col_idx + 1, f'{curr}_MuaTM')
        ws_rate.cell(1, col_idx + 2, f'{curr}_MuaCK')
        ws_rate.cell(1, col_idx + 3, f'{curr}_Ban')
        col_idx += 4
    
    # Data rows: mỗi ngày 1 hàng (sắp xếp tăng dần để MATCH loại 1 hoạt động chính xác)
    unique_dates_asc = sorted(unique_dates)
    for row_i, date_key in enumerate(unique_dates_asc, 2):
        dt = datetime.strptime(date_key, '%Y-%m-%d')
        display_d = dt.strftime('%d/%m/%Y')
        ws_rate.cell(row_i, 1, display_d)
        ws_rate.cell(row_i, 2, f'=IF(ISNUMBER(A{row_i}), A{row_i}, DATE(RIGHT(A{row_i},4), MID(A{row_i},4,2), LEFT(A{row_i},2)))')
        
        df_day = df_work[df_work['_date_key'] == date_key]
        for _, row_data in df_day.iterrows():
            curr = row_data['Mã Ngoại Tệ']
            if curr in currency_col_map:
                c = currency_col_map[curr]
                name_val = row_data.get('Tên Ngoại Tệ', '')
                ws_rate.cell(row_i, c, name_val if pd.notna(name_val) else '')
                ws_rate.cell(row_i, c + 1, row_data['Mua Tiền Mặt'] if pd.notna(row_data['Mua Tiền Mặt']) else None)
                ws_rate.cell(row_i, c + 2, row_data['Mua Chuyển Khoản'] if pd.notna(row_data['Mua Chuyển Khoản']) else None)
                ws_rate.cell(row_i, c + 3, row_data['Bán'] if pd.notna(row_data['Bán']) else None)
    
    rate_max_row = 1 + len(unique_dates)
    rate_max_col = 2 + len(all_currencies) * 4
    
    # ═══════════════════════════════════════════════════════
    # SHEET Dashboard: Date picker + Formula-based table
    # ═══════════════════════════════════════════════════════
    if 'Dashboard' in wb.sheetnames:
        del wb['Dashboard']
    ws = wb.create_sheet('Dashboard', 0)
    ws.sheet_view.showGridLines = False
    
    # ── VCB Color Palette ──
    VCB_GREEN = '00703C'
    VCB_GREEN_LIGHT = 'E8F5E9'
    VCB_WHITE = 'FFFFFF'
    VCB_DARK = '333333'
    VCB_GRAY = '666666'
    VCB_SELL_RED = 'C0392B'
    
    # ── Style definitions ──
    title_font = Font(bold=True, size=22, color=VCB_DARK, name='Arial')
    subtitle_font = Font(italic=True, size=11, color=VCB_GRAY, name='Arial')
    header_fill = PatternFill(start_color=VCB_GREEN, end_color=VCB_GREEN, fill_type='solid')
    header_font = Font(bold=True, color=VCB_WHITE, size=11, name='Arial')
    even_fill = PatternFill(start_color=VCB_GREEN_LIGHT, end_color=VCB_GREEN_LIGHT, fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='C8E6C9'),
        right=Side(style='thin', color='C8E6C9'),
        top=Side(style='thin', color='C8E6C9'),
        bottom=Side(style='thin', color='C8E6C9'),
    )
    date_picker_border = Border(
        left=Side(style='medium', color=VCB_GREEN),
        right=Side(style='medium', color=VCB_GREEN),
        top=Side(style='medium', color=VCB_GREEN),
        bottom=Side(style='medium', color=VCB_GREEN),
    )
    
    # ── Set column widths (mở rộng cho cột so sánh) ──
    col_widths = {
        'A': 3, 'B': 16, 'C': 28, 'D': 20, 'E': 22, 'F': 18,
        'G': 18, 'H': 14, 'I': 18, 'J': 16, 'K': 14, 'L': 3
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    
    # ═══════════════════════════════════════════════════════
    # PHẦN 1: Tiêu đề + Date Pickers (B4 = ngày xem, E4 = ngày so sánh)
    # ═══════════════════════════════════════════════════════
    ws.merge_cells('B2:K2')
    ws['B2'].value = 'Tỷ giá ngoại tệ'
    ws['B2'].font = title_font
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 40
    
    # ── DATE PICKER 1: Ngày xem (B3-B4) ──
    ws['B3'].value = 'Ngày xem'
    ws['B3'].font = Font(size=10, color=VCB_GRAY, name='Arial')
    ws['B3'].alignment = Alignment(horizontal='left')
    
    ws['B4'].value = latest_display
    ws['B4'].font = Font(size=13, color=VCB_DARK, name='Arial', bold=True)
    ws['B4'].alignment = Alignment(horizontal='left', vertical='center')
    ws['B4'].border = date_picker_border
    ws['B4'].fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
    ws.row_dimensions[4].height = 32
    
    # ── Write lists to hidden columns (AA and AB) ──
    # Column 27 (AA) = Dates
    # Column 28 (AB) = Currencies
    for i, d_str in enumerate(display_dates, 1):
        ws.cell(i, 27, d_str)
    
    for i, curr_code in enumerate(all_currencies, 1):
        ws.cell(i, 28, curr_code)
        
    date_range_ref = f"=$AA$1:$AA${len(display_dates)}"
    curr_list_ref = f"=$AB$1:$AB${len(all_currencies)}"
    
    ws.column_dimensions['AA'].hidden = True
    ws.column_dimensions['AB'].hidden = True
    
    ws['C4'].value = '📅 ← Chọn ngày'
    ws['C4'].font = Font(size=10, color=VCB_GRAY, italic=True, name='Arial')
    ws['C4'].alignment = Alignment(horizontal='left', vertical='center')
    
    dv = DataValidation(type="list", formula1=date_range_ref, allow_blank=False)
    dv.prompt = "Chọn ngày để xem tỷ giá"
    dv.promptTitle = "Chọn ngày"
    dv.error = "Vui lòng chọn ngày từ danh sách"
    dv.errorTitle = "Ngày không hợp lệ"
    dv.showInputMessage = True
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add('B4')
    
    # ── DATE PICKER 2: Ngày so sánh (E3-E4) ──
    ws['E3'].value = 'So sánh với ngày'
    ws['E3'].font = Font(size=10, color=VCB_GRAY, name='Arial')
    ws['E3'].alignment = Alignment(horizontal='left')
    
    # Default: ngày cũ nhất gần nhất (ngày hôm qua hoặc ngày trước đó)
    compare_display = display_dates[1] if len(display_dates) > 1 else latest_display
    ws['E4'].value = compare_display
    ws['E4'].font = Font(size=13, color='1565C0', name='Arial', bold=True)
    ws['E4'].alignment = Alignment(horizontal='left', vertical='center')
    ws['E4'].border = Border(
        left=Side(style='medium', color='1565C0'),
        right=Side(style='medium', color='1565C0'),
        top=Side(style='medium', color='1565C0'),
        bottom=Side(style='medium', color='1565C0'),
    )
    ws['E4'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    
    ws['F4'].value = '📅 ← So sánh'
    ws['F4'].font = Font(size=10, color=VCB_GRAY, italic=True, name='Arial')
    ws['F4'].alignment = Alignment(horizontal='left', vertical='center')
    
    dv2 = DataValidation(type="list", formula1=date_range_ref, allow_blank=False)
    dv2.prompt = "Chọn ngày để so sánh"
    dv2.promptTitle = "Ngày so sánh"
    dv2.showInputMessage = True
    dv2.showErrorMessage = True
    ws.add_data_validation(dv2)
    dv2.add('E4')
    
    # ═══════════════════════════════════════════════════════
    # PHẦN 2: Bảng tỷ giá + So sánh 2 ngày + % Thay đổi
    # B=Mã, C=Tên, D=Mua TM, E=Mua CK, F=Bán, G=Bình quân,
    # H=% Thay đổi, I=Bán(SS), J=Chênh lệch, K=% SS
    # ═══════════════════════════════════════════════════════
    table_start = 6
    table_headers = ['Mã ngoại tệ', 'Tên ngoại tệ', 'Mua tiền mặt', 'Mua chuyển khoản',
                     'Bán', 'Bình quân', '% Thay đổi', 'Bán (SS)', 'Chênh lệch', '% So sánh']
    table_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    
    # Color for comparison columns
    compare_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    compare_header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    
    # Header row
    for col_letter, header_text in zip(table_cols, table_headers):
        cell = ws[f'{col_letter}{table_start}']
        cell.value = header_text
        cell.font = header_font
        cell.fill = compare_header_fill if col_letter in ('I', 'J', 'K') else header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[table_start].height = 36
    
    # Data rows
    for i, curr in enumerate(all_currencies):
        row_idx = table_start + 1 + i
        ws.row_dimensions[row_idx].height = 30
        
        # Cột B: Mã ngoại tệ
        cell_code = ws[f'B{row_idx}']
        cell_code.value = curr
        cell_code.font = Font(bold=True, size=11, color=VCB_DARK, name='Arial')
        cell_code.alignment = Alignment(horizontal='center', vertical='center')
        cell_code.border = thin_border
        
        rate_col = currency_col_map[curr]
        rc_name = get_column_letter(rate_col)
        rc_buy = get_column_letter(rate_col + 1)
        rc_ck = get_column_letter(rate_col + 2)
        rc_sell = get_column_letter(rate_col + 3)
        
        match_main = f'MATCH($B$4,\'_RateData\'!$A$2:$A${rate_max_row},0)'
        match_comp = f'MATCH($E$4,\'_RateData\'!$A$2:$A${rate_max_row},0)'
        
        # Cột C: Tên ngoại tệ
        ws[f'C{row_idx}'].value = f'=IFERROR(INDEX(\'_RateData\'!${rc_name}$2:${rc_name}${rate_max_row},{match_main}),"-")'
        ws[f'C{row_idx}'].font = Font(size=11, color=VCB_DARK, name='Arial')
        ws[f'C{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'C{row_idx}'].border = thin_border
        
        # Cột D: Mua tiền mặt
        ws[f'D{row_idx}'].value = f'=IFERROR(INDEX(\'_RateData\'!${rc_buy}$2:${rc_buy}${rate_max_row},{match_main}),"-")'
        ws[f'D{row_idx}'].font = Font(size=11, color=VCB_DARK, name='Arial')
        ws[f'D{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'D{row_idx}'].border = thin_border
        ws[f'D{row_idx}'].number_format = '#,##0.00'
        
        # Cột E: Mua chuyển khoản
        ws[f'E{row_idx}'].value = f'=IFERROR(INDEX(\'_RateData\'!${rc_ck}$2:${rc_ck}${rate_max_row},{match_main}),"-")'
        ws[f'E{row_idx}'].font = Font(size=11, color=VCB_DARK, name='Arial')
        ws[f'E{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'E{row_idx}'].border = thin_border
        ws[f'E{row_idx}'].number_format = '#,##0.00'
        
        # Cột F: Bán
        ws[f'F{row_idx}'].value = f'=IFERROR(INDEX(\'_RateData\'!${rc_sell}$2:${rc_sell}${rate_max_row},{match_main}),"-")'
        ws[f'F{row_idx}'].font = Font(bold=True, size=11, color=VCB_SELL_RED, name='Arial')
        ws[f'F{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'F{row_idx}'].border = thin_border
        ws[f'F{row_idx}'].number_format = '#,##0.00'
        
        # Cột G: Bình quân
        ws[f'G{row_idx}'].value = f'=IFERROR(AVERAGE(E{row_idx},F{row_idx}),"-")'
        ws[f'G{row_idx}'].font = Font(bold=True, size=11, color=VCB_DARK, name='Arial')
        ws[f'G{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'G{row_idx}'].border = thin_border
        ws[f'G{row_idx}'].number_format = '#,##0.00'
        
        # Cột H: % Thay đổi so với ngày so sánh = (F - I) / I * 100
        ws[f'H{row_idx}'].value = f'=IFERROR(IF(AND(ISNUMBER(F{row_idx}),ISNUMBER(I{row_idx}),I{row_idx}<>0),(F{row_idx}-I{row_idx})/I{row_idx}*100,"-"),"-")'
        ws[f'H{row_idx}'].font = Font(bold=True, size=11, name='Arial')
        ws[f'H{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'H{row_idx}'].border = thin_border
        ws[f'H{row_idx}'].number_format = '0.00"%"'
        
        # Cột I: Bán ngày so sánh
        ws[f'I{row_idx}'].value = f'=IFERROR(INDEX(\'_RateData\'!${rc_sell}$2:${rc_sell}${rate_max_row},{match_comp}),"-")'
        ws[f'I{row_idx}'].font = Font(size=11, color='1565C0', name='Arial')
        ws[f'I{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'I{row_idx}'].border = thin_border
        ws[f'I{row_idx}'].number_format = '#,##0.00'
        ws[f'I{row_idx}'].fill = compare_fill
        
        # Cột J: Chênh lệch = F - I
        ws[f'J{row_idx}'].value = f'=IFERROR(IF(AND(ISNUMBER(F{row_idx}),ISNUMBER(I{row_idx})),F{row_idx}-I{row_idx},"-"),"-")'
        ws[f'J{row_idx}'].font = Font(bold=True, size=11, name='Arial')
        ws[f'J{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'J{row_idx}'].border = thin_border
        ws[f'J{row_idx}'].number_format = '+#,##0.00;-#,##0.00;0.00'
        ws[f'J{row_idx}'].fill = compare_fill
        
        # Cột K: % So sánh
        ws[f'K{row_idx}'].value = f'=IFERROR(IF(AND(ISNUMBER(F{row_idx}),ISNUMBER(I{row_idx}),I{row_idx}<>0),(F{row_idx}-I{row_idx})/I{row_idx}*100,"-"),"-")'
        ws[f'K{row_idx}'].font = Font(bold=True, size=11, name='Arial')
        ws[f'K{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'K{row_idx}'].border = thin_border
        ws[f'K{row_idx}'].number_format = '+0.00"%";-0.00"%";0.00"%"'
        ws[f'K{row_idx}'].fill = compare_fill
        
        # Alternating row color
        if i % 2 == 1:
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{col}{row_idx}'].fill = even_fill
            for col in ['I', 'J', 'K']:
                ws[f'{col}{row_idx}'].fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
    
    table_end = table_start + len(all_currencies)
    
    # ── Conditional Formatting: xanh/đỏ cho cột % thay đổi ──
    first_data_row = table_start + 1
    last_data_row = table_end
    
    green_font = Font(bold=True, size=11, color='27AE60', name='Arial')
    red_font = Font(bold=True, size=11, color='C0392B', name='Arial')
    strong_red_font = Font(bold=True, size=12, color='FFFFFF', name='Arial')
    strong_green_font = Font(bold=True, size=12, color='FFFFFF', name='Arial')
    strong_red_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    strong_green_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    
    # Áp dụng cho cột H (% Thay đổi) và K (% So sánh)
    for pct_col in ['H', 'K']:
        pct_range = f'{pct_col}{first_data_row}:{pct_col}{last_data_row}'
        
        # Biến động mạnh > 0.5% (xanh nền)
        ws.conditional_formatting.add(pct_range, FormulaRule(
            formula=[f'{pct_col}{first_data_row}>0.5'],
            fill=strong_green_fill, font=strong_green_font
        ))
        # Biến động mạnh < -0.5% (đỏ nền)
        ws.conditional_formatting.add(pct_range, FormulaRule(
            formula=[f'{pct_col}{first_data_row}<-0.5'],
            fill=strong_red_fill, font=strong_red_font
        ))
        # Tăng nhẹ 0 < x <= 0.5 (font xanh)
        ws.conditional_formatting.add(pct_range, CellIsRule(
            operator='greaterThan', formula=['0'], font=green_font
        ))
        # Giảm nhẹ -0.5 <= x < 0 (font đỏ)
        ws.conditional_formatting.add(pct_range, CellIsRule(
            operator='lessThan', formula=['0'], font=red_font
        ))
    
    # Cột J (Chênh lệch): xanh/đỏ theo dấu
    diff_range = f'J{first_data_row}:J{last_data_row}'
    ws.conditional_formatting.add(diff_range, CellIsRule(
        operator='greaterThan', formula=['0'], font=green_font
    ))
    ws.conditional_formatting.add(diff_range, CellIsRule(
        operator='lessThan', formula=['0'], font=red_font
    ))
    
    # ═══════════════════════════════════════════════════════
    # PHẦN 3: Biểu đồ với Currency Selector
    # ═══════════════════════════════════════════════════════
    chart_section_start = table_end + 3
    
    # Currency selector label + dropdown
    ws[f'B{chart_section_start}'].value = '📈 Biểu đồ biến động tỷ giá (TB tháng)'
    ws[f'B{chart_section_start}'].font = Font(bold=True, size=14, color=VCB_DARK, name='Arial')
    ws.merge_cells(f'B{chart_section_start}:F{chart_section_start}')
    
    curr_selector_row = chart_section_start + 1
    ws[f'B{curr_selector_row}'].value = 'Chọn ngoại tệ:'
    ws[f'B{curr_selector_row}'].font = Font(size=11, color=VCB_GRAY, name='Arial')
    ws[f'B{curr_selector_row}'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws[f'C{curr_selector_row}'].value = 'USD'
    ws[f'C{curr_selector_row}'].font = Font(size=13, color=VCB_DARK, name='Arial', bold=True)
    ws[f'C{curr_selector_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'C{curr_selector_row}'].border = date_picker_border
    ws[f'C{curr_selector_row}'].fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
    
    # Data Validation cho currency selector
    dv_curr = DataValidation(type="list", formula1=curr_list_ref, allow_blank=False)
    dv_curr.prompt = "Chọn ngoại tệ để xem biểu đồ"
    dv_curr.promptTitle = "Chọn ngoại tệ"
    dv_curr.showInputMessage = True
    dv_curr.showErrorMessage = True
    ws.add_data_validation(dv_curr)
    dv_curr.add(f'C{curr_selector_row}')
    
    # ── Sheet _ChartData: monthly averages cho TẤT CẢ ngoại tệ ──
    if '_ChartData' in wb.sheetnames:
        del wb['_ChartData']
    ws_chart_data = wb.create_sheet('_ChartData')
    ws_chart_data.sheet_state = 'hidden'
    
    df_chart = df.copy()
    df_chart['Ngày'] = pd.to_datetime(df_chart['Ngày Cập Nhật'], format='mixed')
    df_chart['Tháng'] = df_chart['Ngày'].dt.to_period('M')
    
    # Build monthly averages for ALL currencies
    periods = sorted(df_chart['Tháng'].unique())
    
    # Header: A=Tháng, then each currency gets 1 column (Bán TB)
    ws_chart_data.cell(1, 1, 'Tháng')
    chart_curr_col = {}
    for ci, c in enumerate(all_currencies, 2):
        ws_chart_data.cell(1, ci, c)
        chart_curr_col[c] = ci
    
    for pi, period in enumerate(periods, 2):
        ws_chart_data.cell(pi, 1, str(period))
        grp = df_chart[df_chart['Tháng'] == period]
        for c in all_currencies:
            curr_data = grp[grp['Mã Ngoại Tệ'] == c]
            sell_avg = curr_data['Bán'].mean() if not curr_data['Bán'].isna().all() else None
            ws_chart_data.cell(pi, chart_curr_col[c], round(sell_avg, 2) if sell_avg else None)
    
    chart_data_max_row = 1 + len(periods)
    
    # ── Mini table trên Dashboard: dùng INDEX/MATCH theo currency selector ──
    mini_start = curr_selector_row + 2
    mini_headers = ['Tháng', 'Giá Bán (TB)']
    mini_header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    
    ws.cell(mini_start, 2, 'Tháng')
    ws.cell(mini_start, 2).font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    ws.cell(mini_start, 2).fill = mini_header_fill
    ws.cell(mini_start, 2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(mini_start, 2).border = thin_border
    
    ws.cell(mini_start, 3, 'Giá Bán (TB)')
    ws.cell(mini_start, 3).font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    ws.cell(mini_start, 3).fill = mini_header_fill
    ws.cell(mini_start, 3).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(mini_start, 3).border = thin_border
    
    for pi, period in enumerate(periods):
        row_idx = mini_start + 1 + pi
        # Cột B: Tháng (static)
        ws.cell(row_idx, 2, str(period))
        ws.cell(row_idx, 2).border = thin_border
        ws.cell(row_idx, 2).font = Font(size=10, name='Arial')
        ws.cell(row_idx, 2).alignment = Alignment(horizontal='center')
        
        # Cột C: Giá bán TB = INDEX(_ChartData, row, MATCH(currency_selector, _ChartData header, 0))
        ws.cell(row_idx, 3).value = f'=IFERROR(INDEX(\'_ChartData\'!$A${pi+2}:${get_column_letter(1+len(all_currencies))}${pi+2},1,MATCH($C${curr_selector_row},\'_ChartData\'!$B$1:${get_column_letter(1+len(all_currencies))}$1,0)),"-")'
        ws.cell(row_idx, 3).border = thin_border
        ws.cell(row_idx, 3).font = Font(size=10, name='Arial')
        ws.cell(row_idx, 3).alignment = Alignment(horizontal='right')
        ws.cell(row_idx, 3).number_format = '#,##0.00'
        
        if pi % 2 == 0:
            ws.cell(row_idx, 2).fill = PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid')
            ws.cell(row_idx, 3).fill = PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid')
    
    mini_end = mini_start + len(periods)
    
    # Line Chart
    chart = LineChart()
    chart.title = "Biến Động Tỷ Giá Bán (TB Tháng)"
    chart.style = 10
    chart.y_axis.title = "VND"
    chart.x_axis.title = "Tháng"
    chart.width = 28
    chart.height = 14
    chart.legend.position = 'b'
    
    cats = Reference(ws, min_col=2, min_row=mini_start + 1, max_row=mini_end)
    vals = Reference(ws, min_col=3, min_row=mini_start, max_row=mini_end)
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    
    s = chart.series[0]
    s.graphicalProperties.line.width = 25000
    s.graphicalProperties.line.solidFill = '2E7D32'
    s.smooth = True
    
    chart_row = mini_end + 2
    ws.add_chart(chart, f'B{chart_row}')
    
    # ═══════════════════════════════════════════════════════
    # PHẦN 4: Liên kết hữu ích
    # ═══════════════════════════════════════════════════════
    links_row = chart_row + 16
    ws.merge_cells(f'B{links_row}:K{links_row}')
    ws[f'B{links_row}'].value = "🔗 Liên Kết Phân Tích Tỷ Giá USD/VND"
    ws[f'B{links_row}'].font = Font(bold=True, size=12, color=VCB_DARK, name='Arial')
    
    links = [
        ("Investing.com - USD/VND Phân tích kỹ thuật", "https://vn.investing.com/currencies/usd-vnd-technical"),
        ("Vietstock - Tin tức Ngoại Tệ chuyên sâu", "https://vietstock.vn/tien-te.htm"),
        ("TradingView - Biểu đồ USD/VND", "https://vn.tradingview.com/symbols/USDVND/"),
    ]
    
    for i, (text, url) in enumerate(links, 1):
        cell_ref = f'B{links_row + i}'
        ws.merge_cells(f'B{links_row + i}:K{links_row + i}')
        cell = ws[cell_ref]
        cell.value = text
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single", size=11, name='Arial')
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Source footer
    source_row = links_row + len(links) + 2
    ws.merge_cells(f'B{source_row}:K{source_row}')
    ws[f'B{source_row}'].value = "Nguồn: Ngân hàng TMCP Ngoại thương Việt Nam (Vietcombank)"
    ws[f'B{source_row}'].font = Font(italic=True, size=9, color='999999', name='Arial')
    ws[f'B{source_row}'].alignment = Alignment(horizontal='center')
    
    # ── Tạo sheet TheoDoi_USD để báo cáo riêng USD ──
    _create_usd_monthly_sheet(wb)
    
    print(f"📈 Dashboard nâng cao: {len(all_currencies)} ngoại tệ, {len(unique_dates)} ngày, {len(periods)} tháng, so sánh 2 ngày")


def _create_usd_monthly_sheet(wb):
    """Tạo sheet TheoDoi_Thang_USD để theo dõi tỷ giá USD theo tháng chọn."""
    sheet_name = 'TheoDoi_Thang_USD'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    # Chèn vào vị trí thứ 3 (chỉ sau Dashboard và TheoDoi_USD)
    ws = wb.create_sheet(sheet_name, 2)
    ws.sheet_view.showGridLines = False
    
    # ── VCB Color Palette ──
    VCB_GREEN = '00703C'
    VCB_GREEN_LIGHT = 'E8F5E9'
    VCB_WHITE = 'FFFFFF'
    VCB_DARK = '333333'
    VCB_GRAY = '666666'
    
    # Style definitions
    title_font = Font(bold=True, size=18, color=VCB_DARK, name='Arial')
    header_fill = PatternFill(start_color=VCB_GREEN, end_color=VCB_GREEN, fill_type='solid')
    header_font = Font(bold=True, color=VCB_WHITE, size=11, name='Arial')
    even_fill = PatternFill(start_color='F5F9F6', end_color='F5F9F6', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC'),
    )
    
    # Column widths
    col_widths = {
        'A': 3, 'B': 6, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 18, 'H': 18, 'I': 3
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
        
    # Title
    ws.merge_cells('B2:H2')
    ws['B2'].value = 'BÁO CÁO TỶ GIÁ USD THEO THÁNG'
    ws['B2'].font = title_font
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 35
    
    # Selectors Year, Month and Bank
    ws['B4'].value = 'Chọn Năm:'
    ws['B4'].font = Font(size=11, color=VCB_GRAY, name='Arial', bold=True)
    ws['B4'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws['C4'].value = 2026
    ws['C4'].font = Font(size=12, color=VCB_DARK, name='Arial', bold=True)
    ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C4'].border = thin_border
    ws['C4'].fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    ws['E4'].value = 'Chọn Tháng:'
    ws['E4'].font = Font(size=11, color=VCB_GRAY, name='Arial', bold=True)
    ws['E4'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws['F4'].value = 6  # default June
    ws['F4'].font = Font(size=12, color=VCB_DARK, name='Arial', bold=True)
    ws['F4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F4'].border = Border(
        left=Side(style='medium', color=VCB_GREEN),
        right=Side(style='medium', color=VCB_GREEN),
        top=Side(style='medium', color=VCB_GREEN),
        bottom=Side(style='medium', color=VCB_GREEN),
    )
    ws['F4'].fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
    
    ws['G4'].value = 'Ngân Hàng:'
    ws['G4'].font = Font(size=11, color=VCB_GRAY, name='Arial', bold=True)
    ws['G4'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws['H4'].value = 'Vietcombank'  # default Bank
    ws['H4'].font = Font(size=12, color=VCB_DARK, name='Arial', bold=True)
    ws['H4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H4'].border = Border(
        left=Side(style='medium', color=VCB_GREEN),
        right=Side(style='medium', color=VCB_GREEN),
        top=Side(style='medium', color=VCB_GREEN),
        bottom=Side(style='medium', color=VCB_GREEN),
    )
    ws['H4'].fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
    
    ws.row_dimensions[4].height = 28
    
    # Write month list to hidden column AA
    for m in range(1, 13):
        ws.cell(m, 27, m)
    ws.column_dimensions['AA'].hidden = True
    
    # Write bank list to hidden column AB
    banks = ['Vietcombank', 'VietinBank', 'BIDV', 'Techcombank', 'ACB', 'SeaBank']
    for i, b in enumerate(banks, 1):
        ws.cell(i, 28, b)
    ws.column_dimensions['AB'].hidden = True
    
    # Dropdown validation for Month
    dv_month = DataValidation(type="list", formula1="=$AA$1:$AA$12", allow_blank=False)
    dv_month.prompt = "Chọn tháng cần theo dõi (1 - 12)"
    dv_month.promptTitle = "Chọn tháng"
    ws.add_data_validation(dv_month)
    dv_month.add(ws['F4'])
    
    # Dropdown validation for Bank
    dv_bank = DataValidation(type="list", formula1="=$AB$1:$AB$6", allow_blank=False)
    dv_bank.prompt = "Chọn ngân hàng cần xem tỷ giá"
    dv_bank.promptTitle = "Chọn ngân hàng"
    ws.add_data_validation(dv_bank)
    dv_bank.add(ws['H4'])
    
    # Table headers
    headers = ['STT', 'Ngày theo lịch', 'Ngày áp dụng', 'Mua tiền mặt', 'Mua chuyển khoản', 'Bán', 'Bình quân']
    cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
    for col_letter, text in zip(cols, headers):
        cell = ws[f'{col_letter}6']
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[6].height = 32
    
    # Data rows
    for d in range(1, 32):
        row_idx = 6 + d
        ws.row_dimensions[row_idx].height = 26
        
        # B: STT
        ws[f'B{row_idx}'].value = f'=IF(C{row_idx}<>"", {d}, "")'
        ws[f'B{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{row_idx}'].font = Font(size=10, name='Arial')
        ws[f'B{row_idx}'].border = thin_border
        
        # C: Ngày lịch
        ws[f'C{row_idx}'].value = f'=IF(MONTH(DATE($C$4, $F$4, {d}))=$F$4, DATE($C$4, $F$4, {d}), "")'
        ws[f'C{row_idx}'].number_format = 'dd/mm/yyyy'
        ws[f'C{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'C{row_idx}'].font = Font(size=11, name='Arial')
        ws[f'C{row_idx}'].border = thin_border
        
        # D: Ngày áp dụng
        ws[f'D{row_idx}'].value = f'=IF(C{row_idx}<>"", IF(E{row_idx}<>"-", C{row_idx}, "-"), "")'
        ws[f'D{row_idx}'].number_format = 'dd/mm/yyyy'
        ws[f'D{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'D{row_idx}'].font = Font(size=11, name='Arial')
        ws[f'D{row_idx}'].border = thin_border
        
        # E: Mua tiền mặt
        ws[f'E{row_idx}'].value = f'=IF(C{row_idx}<>"", IF(SUMIFS(Data_TheoDoi_USD!$C$2:$C$10000, Data_TheoDoi_USD!$A$2:$A$10000, TEXT(C{row_idx}, "dd/mm/yyyy"), Data_TheoDoi_USD!$B$2:$B$10000, $H$4)=0, "-", SUMIFS(Data_TheoDoi_USD!$C$2:$C$10000, Data_TheoDoi_USD!$A$2:$A$10000, TEXT(C{row_idx}, "dd/mm/yyyy"), Data_TheoDoi_USD!$B$2:$B$10000, $H$4)), "")'
        ws[f'E{row_idx}'].number_format = '#,##0.00'
        ws[f'E{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'E{row_idx}'].font = Font(size=11, name='Arial')
        ws[f'E{row_idx}'].border = thin_border
        
        # F: Mua chuyển khoản
        ws[f'F{row_idx}'].value = f'=IF(C{row_idx}<>"", IF(SUMIFS(Data_TheoDoi_USD!$D$2:$D$10000, Data_TheoDoi_USD!$A$2:$A$10000, TEXT(C{row_idx}, "dd/mm/yyyy"), Data_TheoDoi_USD!$B$2:$B$10000, $H$4)=0, "-", SUMIFS(Data_TheoDoi_USD!$D$2:$D$10000, Data_TheoDoi_USD!$A$2:$A$10000, TEXT(C{row_idx}, "dd/mm/yyyy"), Data_TheoDoi_USD!$B$2:$B$10000, $H$4)), "")'
        ws[f'F{row_idx}'].number_format = '#,##0.00'
        ws[f'F{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'F{row_idx}'].font = Font(size=11, name='Arial')
        ws[f'F{row_idx}'].border = thin_border
        
        # G: Bán
        ws[f'G{row_idx}'].value = f'=IF(C{row_idx}<>"", IF(SUMIFS(Data_TheoDoi_USD!$E$2:$E$10000, Data_TheoDoi_USD!$A$2:$A$10000, TEXT(C{row_idx}, "dd/mm/yyyy"), Data_TheoDoi_USD!$B$2:$B$10000, $H$4)=0, "-", SUMIFS(Data_TheoDoi_USD!$E$2:$E$10000, Data_TheoDoi_USD!$A$2:$A$10000, TEXT(C{row_idx}, "dd/mm/yyyy"), Data_TheoDoi_USD!$B$2:$B$10000, $H$4)), "")'
        ws[f'G{row_idx}'].number_format = '#,##0.00'
        ws[f'G{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'G{row_idx}'].font = Font(size=11, name='Arial')
        ws[f'G{row_idx}'].border = thin_border
        
        # H: Bình quân
        ws[f'H{row_idx}'].value = f'=IF(AND(C{row_idx}<>"", E{row_idx}<>"-", F{row_idx}<>"-", G{row_idx}<>"-"), IFERROR(AVERAGE(F{row_idx}, G{row_idx}), "-"), "")'
        ws[f'H{row_idx}'].number_format = '#,##0.00'
        ws[f'H{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'H{row_idx}'].font = Font(bold=True, size=11, name='Arial', color='0E6655')
        ws[f'H{row_idx}'].border = thin_border
        
        # Alternating zebra rows
        if d % 2 == 0:
            for col in cols:
                ws[f'{col}{row_idx}'].fill = even_fill
                
    # Row 38: Monthly Average
    avg_row_idx = 38
    ws.row_dimensions[avg_row_idx].height = 28
    
    ws[f'C{avg_row_idx}'].value = 'BÌNH QUÂN THÁNG'
    ws[f'C{avg_row_idx}'].font = Font(bold=True, size=11, name='Arial', color=VCB_GREEN)
    ws[f'C{avg_row_idx}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'C{avg_row_idx}'].border = thin_border
    
    ws[f'B{avg_row_idx}'].border = thin_border
    ws[f'D{avg_row_idx}'].border = thin_border
    
    for col in ['E', 'F', 'G', 'H']:
        ws[f'{col}{avg_row_idx}'].value = f'=IFERROR(AVERAGE({col}7:{col}37), "-")'
        ws[f'{col}{avg_row_idx}'].number_format = '#,##0.00'
        ws[f'{col}{avg_row_idx}'].font = Font(bold=True, size=11, name='Arial', color=VCB_GREEN)
        ws[f'{col}{avg_row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'{col}{avg_row_idx}'].border = thin_border


def get_exchange_rates_json_fallback():
    """Fallback to get VCB exchange rates from the modern JSON API if the XML portal fails."""
    logger.info("Thử lấy tỷ giá VCB qua API JSON fallback...")
    url = 'https://www.vietcombank.com.vn/api/exchangerates'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36'
    }
    date_str = datetime.now().strftime('%Y-%m-%d')
    try:
        resp = requests.get(url, params={'date': date_str}, headers=headers, timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            update_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            rates = []
            for item in data.get('Data', []):
                currency_code = item.get('currencyCode', '').strip()
                currency_name = item.get('currencyName', '').strip()
                buy = item.get('cash')
                transfer = item.get('transfer')
                sell = item.get('sell')
                
                rates.append({
                    'Ngày Cập Nhật': update_time_str,
                    'Mã Ngoại Tệ': currency_code,
                    'Tên Ngoại Tệ': currency_name,
                    'Mua Tiền Mặt': _to_float(buy),
                    'Mua Chuyển Khoản': _to_float(transfer),
                    'Bán': _to_float(sell)
                })
            if rates:
                logger.info(f"Lấy thành công {len(rates)} ngoại tệ qua API JSON.")
                return rates
    except Exception as e:
        logger.error(f"Lỗi khi gọi API JSON fallback: {e}")
    return None


def get_exchange_rates_with_retry():
    """Fetch rates with retry logic and JSON fallback."""
    for attempt in range(1, RETRY_COUNT + 1):
        logger.info(f"Lần thử {attempt}/{RETRY_COUNT}...")
        rates = get_exchange_rates()
        if rates:
            logger.info(f"Lấy được {len(rates)} ngoại tệ.")
            return rates
        if attempt < RETRY_COUNT:
            logger.warning(f"Thất bại lần {attempt}. Thử lại sau {RETRY_DELAY}s...")
            time.sleep(RETRY_DELAY)
            
    logger.warning(f"Thất bại sau {RETRY_COUNT} lần thử qua XML. Đang chuyển sang API JSON fallback...")
    rates = get_exchange_rates_json_fallback()
    if rates:
        return rates
        
    logger.error("KHÔNG thể lấy dữ liệu tỷ giá từ cả hai nguồn!")
    return None


if __name__ == "__main__":
    logger.info("=" * 50)
    logger.info("Bắt đầu cập nhật tỷ giá VCB")
    
    rates = get_exchange_rates_with_retry()
    if rates:
        save_to_excel(rates, OUTPUT_FILE)
        logger.info("Hoàn tất cập nhật tỷ giá.")
        send_notification("Tỷ Giá VCB", f"Cập nhật thành công {len(rates)} ngoại tệ")
    else:
        logger.error("KHÔNG thể lấy dữ liệu tỷ giá!")
        send_notification("Tỷ Giá VCB - LỖI", "Không thể lấy dữ liệu tỷ giá VCB!")
