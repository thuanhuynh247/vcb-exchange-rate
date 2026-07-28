"""
Fetch VCB exchange rates from 2026-01-01 to present.
Step 1: Fetch all data and save raw CSV
Step 2: Build Excel with Dashboard using get_rates.py functions
"""
import requests
import pandas as pd
from datetime import datetime, timedelta
import concurrent.futures
import os, sys

API_URL = 'https://www.vietcombank.com.vn/api/exchangerates'
OUTPUT_FILE = r'D:\Tygia-Tudong\TyGia_Banking.xlsx'
CSV_FILE = r'D:\Tygia-Tudong\vcb_2026_raw.csv'
HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

CURRENCY_NAMES = {
    'USD': 'US DOLLAR', 'EUR': 'EURO', 'GBP': 'POUND STERLING',
    'JPY': 'YEN', 'AUD': 'AUSTRALIAN DOLLAR', 'CAD': 'CANADIAN DOLLAR',
    'CHF': 'SWISS FRANC', 'SGD': 'SINGAPORE DOLLAR', 'CNY': 'YUAN RENMINBI',
    'HKD': 'HONGKONG DOLLAR', 'THB': 'THAILAND BAHT', 'KRW': 'KOREAN WON',
    'MYR': 'MALAYSIAN RINGGIT', 'INR': 'INDIAN RUPEE', 'KWD': 'KUWAITI DINAR',
    'SAR': 'SAUDI RIAL', 'NOK': 'NORWEGIAN KRONER', 'SEK': 'SWEDISH KRONA',
    'DKK': 'DANISH KRONE', 'RUB': 'RUSSIAN RUBLE',
}

VIET_COLS = ['Ngày Cập Nhật', 'Mã Ngoại Tệ', 'Tên Ngoại Tệ',
             'Mua Tiền Mặt', 'Mua Chuyển Khoản', 'Bán']


def _to_float(val):
    try:
        v = float(str(val).replace(',', ''))
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None


def fetch(date_str):
    try:
        resp = requests.get(API_URL, params={'date': date_str}, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        api_date = data.get('Date', '')[:10]
        rows = []
        for item in data.get('Data', []):
            code = item.get('currencyCode', '')
            if code:
                rows.append([
                    api_date, code,
                    CURRENCY_NAMES.get(code, item.get('currencyName', '').strip()),
                    _to_float(item.get('cash', '')),
                    _to_float(item.get('transfer', '')),
                    _to_float(item.get('sell', '')),
                ])
        return rows if rows else None
    except:
        return None


def step1_fetch():
    """Fetch all data from 2026-01-01 to yesterday, save as CSV."""
    start = datetime(2026, 1, 1)
    end = datetime.now() - timedelta(days=1)

    dates = []
    d = start
    while d <= end:
        dates.append(d.strftime('%Y-%m-%d'))
        d += timedelta(days=1)

    total = len(dates)
    print(f'[STEP 1] Fetching {total} days: {start.date()} -> {end.date()}')

    all_rows = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
        futs = {ex.submit(fetch, d): d for d in dates}
        for i, f in enumerate(concurrent.futures.as_completed(futs)):
            rows = f.result()
            if rows:
                all_rows.extend(rows)
            if (i + 1) % 20 == 0 or (i + 1) == total:
                print(f'  {i + 1}/{total} done ({len(all_rows)} records)')

    if not all_rows:
        print('[ERROR] No data fetched!')
        return False

    # Save CSV
    df = pd.DataFrame(all_rows, columns=VIET_COLS)
    df['Bình Quân CK & Bán'] = df[['Mua Chuyển Khoản', 'Bán']].mean(axis=1).round(2)
    # Dedup
    df = df.drop_duplicates(subset=['Ngày Cập Nhật', 'Mã Ngoại Tệ'], keep='first')
    df = df.sort_values(['Ngày Cập Nhật', 'Mã Ngoại Tệ'], ascending=[False, True]).reset_index(drop=True)
    df.to_csv(CSV_FILE, index=False, encoding='utf-8-sig')
    print(f'[OK] Saved {len(df)} records to CSV ({df["Ngày Cập Nhật"].nunique()} dates)')
    return True


def step2_build_excel():
    """Build Excel from CSV using get_rates.py Dashboard functions."""
    if not os.path.exists(CSV_FILE):
        print('[ERROR] CSV not found!')
        return

    df = pd.read_csv(CSV_FILE)
    print(f'[STEP 2] Building Excel from {len(df)} records...')

    # Import dashboard functions
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if script_dir not in sys.path:
        sys.path.insert(0, script_dir)
    from get_rates import _format_data_sheet, _create_dashboard_with_date_picker
    from openpyxl import load_workbook

    # Write Data sheet
    print('  Writing Data sheet...')
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data', index=False)

    # Format
    print('  Formatting Data sheet...')
    wb = load_workbook(OUTPUT_FILE)
    _format_data_sheet(wb['Data'])
    wb.save(OUTPUT_FILE)

    # Dashboard
    print('  Creating Dashboard...')
    try:
        wb = load_workbook(OUTPUT_FILE)
        _create_dashboard_with_date_picker(wb, df)
        wb.save(OUTPUT_FILE)
        unique_dates = df['Ngày Cập Nhật'].nunique()
        print(f'[DONE] Excel updated: {len(df)} records, {unique_dates} dates')
        print(f'  File: {OUTPUT_FILE}')
    except Exception as e:
        print(f'[ERROR] Dashboard: {e}')
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    print('=== VCB Exchange Rate: 2026-01-01 to Present ===')
    if step1_fetch():
        step2_build_excel()
    print('\nDone!')
