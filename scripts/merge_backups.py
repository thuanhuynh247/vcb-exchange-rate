"""
Merge the largest backup file (with historical data) into TyGia_Banking.xlsx.
"""
import pandas as pd
import os
import sys

OUTPUT_FILE = r'D:\Tygia-Tudong\TyGia_Banking.xlsx'
BACKUP_DIR = r'D:\Tygia-Tudong\Backup'

EXPECTED_COLS = ['Ngày Cập Nhật', 'Mã Ngoại Tệ', 'Tên Ngoại Tệ',
                 'Mua Tiền Mặt', 'Mua Chuyển Khoản', 'Bán']

FIVE_COLS = ['Ngày Cập Nhật', 'Mã Ngoại Tệ', 'Mua Tiền Mặt', 'Mua Chuyển Khoản', 'Bán']

CURRENCY_NAMES = {
    'USD': 'US DOLLAR', 'EUR': 'EURO', 'GBP': 'POUND STERLING',
    'JPY': 'YEN', 'AUD': 'AUSTRALIAN DOLLAR', 'CAD': 'CANADIAN DOLLAR',
    'CHF': 'SWISS FRANC', 'SGD': 'SINGAPORE DOLLAR', 'CNY': 'YUAN RENMINBI',
    'HKD': 'HONGKONG DOLLAR', 'THB': 'THAILAND BAHT', 'KRW': 'KOREAN WON',
    'MYR': 'MALAYSIAN RINGGIT', 'INR': 'INDIAN RUPEE', 'KWD': 'KUWAITI DINAR',
    'SAR': 'SAUDI RIAL', 'NOK': 'NORWEGIAN KRONER', 'SEK': 'SWEDISH KRONA',
    'DKK': 'DANISH KRONE', 'RUB': 'RUSSIAN RUBLE',
}

# Step 1: Find and read ALL backup files with 'Data' sheet
all_dfs = []

for f in sorted(os.listdir(BACKUP_DIR)):
    if not f.endswith('.xlsx'):
        continue
    fp = os.path.join(BACKUP_DIR, f)
    try:
        xls = pd.ExcelFile(fp)
        sheet = 'Data' if 'Data' in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(fp, sheet_name=sheet)
        
        cols = list(df.columns)
        if cols == EXPECTED_COLS and len(df) > 0:
            # Full 6-column format
            all_dfs.append(df)
            print(f"OK (6col): {f}: {len(df)} rows")
        elif cols == FIVE_COLS and len(df) > 0:
            # Old 5-column format - insert Tên Ngoại Tệ
            df.insert(2, 'Tên Ngoại Tệ', df['Mã Ngoại Tệ'].map(CURRENCY_NAMES).fillna(''))
            all_dfs.append(df)
            print(f"OK (5col): {f}: {len(df)} rows")
        else:
            print(f"SKIP: {f}: {len(cols)} cols, {list(df.columns)[:3]}...")
    except Exception as e:
        print(f"ERR: {f}: {str(e)[:50]}")

# Also try to read current file
try:
    df_cur = pd.read_excel(OUTPUT_FILE, sheet_name='Data')
    if list(df_cur.columns) == EXPECTED_COLS:
        all_dfs.append(df_cur)
        print(f"OK: Current file: {len(df_cur)} rows")
except:
    print("Current file not readable")

if not all_dfs:
    print("ERROR: No data found!")
    sys.exit(1)

# Step 2: Combine & dedup
df_all = pd.concat(all_dfs, ignore_index=True)
print(f"\nCombined: {len(df_all)} rows")

df_all['_dk'] = pd.to_datetime(df_all['Ngày Cập Nhật'], format='mixed', errors='coerce').dt.strftime('%Y-%m-%d')
df_all = df_all.dropna(subset=['_dk'])
df_all = df_all.drop_duplicates(subset=['_dk', 'Mã Ngoại Tệ'], keep='first')
df_all = df_all.sort_values(['_dk', 'Mã Ngoại Tệ'], ascending=[False, True]).reset_index(drop=True)
df_all['Ngày Cập Nhật'] = df_all['_dk']
df_all.drop(columns=['_dk'], inplace=True)

unique_dates = sorted(pd.to_datetime(df_all['Ngày Cập Nhật'], format='mixed').dt.strftime('%Y-%m-%d').unique())
print(f"After dedup: {len(df_all)} rows")
print(f"Date range: {unique_dates[0]} to {unique_dates[-1]}")
print(f"Unique dates: {len(unique_dates)}")

# Step 3: Write Excel + Dashboard
script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

from openpyxl import load_workbook
from get_rates import _format_data_sheet, _create_dashboard_with_date_picker

TEMP_FILE = OUTPUT_FILE + '.tmp.xlsx'

with pd.ExcelWriter(TEMP_FILE, engine='openpyxl') as writer:
    df_all.to_excel(writer, sheet_name='Data', index=False)

wb = load_workbook(TEMP_FILE)
_format_data_sheet(wb['Data'])
_create_dashboard_with_date_picker(wb, df_all)
wb.save(TEMP_FILE)

try:
    if os.path.exists(OUTPUT_FILE):
        os.remove(OUTPUT_FILE)
    os.rename(TEMP_FILE, OUTPUT_FILE)
    print(f"\nSaved to {OUTPUT_FILE}")
except PermissionError:
    print(f"\nFile locked! Saved as: {TEMP_FILE}")
    print("Close Excel, then rename manually.")

print(f"Sheets: {wb.sheetnames}")
