"""
Test script để kiểm tra cấu trúc và format file TyGia_Banking.xlsx
Đồng bộ với get_rates.py mới: Date Picker, so sánh 2 ngày, biểu đồ chọn ngoại tệ,
conditional formatting, retry/logging.
"""
import pandas as pd
from openpyxl import load_workbook
import os

OUTPUT_FILE = r'D:\Tygia-Tudong\TyGia_Banking.xlsx'
LOG_FILE = r'D:\Tygia-Tudong\vcb_rates.log'

def verify_excel():
    passed = 0
    failed = 0

    def check(condition, ok_msg, fail_msg):
        nonlocal passed, failed
        if condition:
            print(f"  ✅ {ok_msg}")
            passed += 1
        else:
            print(f"  ❌ {fail_msg}")
            failed += 1

    # ── 0. File tồn tại ──
    if not os.path.exists(OUTPUT_FILE):
        print("❌ LỖI: File Excel chưa được tạo.")
        return
    print("📂 Đã tìm thấy file Excel.\n")

    # ── 1. Kiểm tra Data sheet bằng pandas ──
    print("═══ KIỂM TRA DATA SHEET ═══")
    try:
        df = pd.read_excel(OUTPUT_FILE, sheet_name='Data')
        expected_columns = ['Ngày Cập Nhật', 'Mã Ngoại Tệ', 'Tên Ngoại Tệ',
                            'Mua Tiền Mặt', 'Mua Chuyển Khoản', 'Bán', 'Bình Quân CK & Bán']

        check(
            list(df.columns) == expected_columns,
            f"Đúng 7 cột: {list(df.columns)}",
            f"Cột không khớp. Mong đợi: {expected_columns}, Hiện tại: {list(df.columns)}"
        )

        check(not df.empty, f"Có {len(df)} dòng dữ liệu.", "File trống!")

        if not df.empty:
            unique_currencies = sorted(df['Mã Ngoại Tệ'].unique())
            n_currencies = len(unique_currencies)
            check(
                n_currencies >= 10,
                f"Có {n_currencies} loại ngoại tệ: {', '.join(unique_currencies[:8])}...",
                f"Chỉ có {n_currencies} ngoại tệ (cần >= 10): {unique_currencies}"
            )

            check(
                'USD' in unique_currencies and 'EUR' in unique_currencies,
                "Có USD và EUR trong dữ liệu.",
                f"Thiếu USD hoặc EUR! Có: {unique_currencies}"
            )

            # Kiểm tra cột Tên Ngoại Tệ có dữ liệu
            has_names = df['Tên Ngoại Tệ'].notna().any()
            check(has_names, "Cột 'Tên Ngoại Tệ' có dữ liệu.", "Cột 'Tên Ngoại Tệ' trống!")

    except Exception as e:
        print(f"  ❌ LỖI đọc Data sheet: {e}")
        failed += 1

    # ── 2. Kiểm tra format bằng openpyxl ──
    print("\n═══ KIỂM TRA FORMAT ═══")
    try:
        wb = load_workbook(OUTPUT_FILE)

        # Dashboard là sheet đầu tiên
        check(
            wb.sheetnames[0] == 'Dashboard',
            f"Dashboard là sheet đầu tiên. Sheets: {wb.sheetnames}",
            f"Dashboard KHÔNG phải sheet đầu tiên. Sheets: {wb.sheetnames}"
        )

        # Kiểm tra Data sheet format
        ws_data = wb['Data']
        check(
            all(cell.font.bold for cell in ws_data[1]),
            "Headers Data sheet đã in đậm.",
            "Headers Data sheet chưa in đậm."
        )

        check(
            all(cell.alignment.horizontal == 'center' for cell in ws_data[1]),
            "Headers Data sheet đã căn giữa.",
            "Headers Data sheet chưa căn giữa."
        )

        # Header background = VCB green (#00703C)
        header_color = ws_data['A1'].fill.start_color.rgb
        check(
            header_color and '00703C' in str(header_color),
            f"Header Data dùng màu VCB green ({header_color}).",
            f"Header Data không đúng màu VCB green (hiện: {header_color})."
        )

        # Cột D (Mua Tiền Mặt) format số
        if ws_data.max_row >= 2:
            cell_d2 = ws_data['D2']
            check(
                cell_d2.number_format == '#,##0.00',
                f"Cột số liệu format đúng (#,##0.00).",
                f"Cột số liệu format sai (hiện: {cell_d2.number_format})."
            )

        # AutoFilter
        check(
            ws_data.auto_filter.ref is not None,
            "AutoFilter đã bật trên Data sheet.",
            "AutoFilter chưa bật!"
        )

        # Freeze panes
        check(
            str(ws_data.freeze_panes) == 'A2',
            "Freeze panes tại A2.",
            f"Freeze panes không đúng (hiện: {ws_data.freeze_panes})."
        )

    except Exception as e:
        print(f"  ❌ LỖI kiểm tra format: {e}")
        failed += 1

    # ── 3. Kiểm tra Dashboard sheet ──
    print("\n═══ KIỂM TRA DASHBOARD ═══")
    try:
        ws_dash = wb['Dashboard']

        # Tiêu đề
        check(
            ws_dash['B2'].value == 'Tỷ giá ngoại tệ',
            "Tiêu đề Dashboard đúng: 'Tỷ giá ngoại tệ'.",
            f"Tiêu đề Dashboard sai: '{ws_dash['B2'].value}'."
        )

        # Header bảng tỷ giá (dòng 6)
        dash_header_color = ws_dash['B6'].fill.start_color.rgb
        check(
            dash_header_color and '00703C' in str(dash_header_color),
            f"Header bảng Dashboard dùng màu VCB green ({dash_header_color}).",
            f"Header bảng Dashboard sai màu (hiện: {dash_header_color})."
        )

        # Kiểm tra có dữ liệu ngoại tệ
        first_currency = ws_dash['B7'].value
        check(
            first_currency is not None and len(str(first_currency)) == 3,
            f"Bảng Dashboard có dữ liệu ngoại tệ (đầu tiên: {first_currency}).",
            f"Bảng Dashboard trống hoặc sai format (B7: {first_currency})."
        )

        # Grid lines ẩn
        check(
            not ws_dash.sheet_view.showGridLines,
            "Dashboard ẩn grid lines.",
            "Dashboard chưa ẩn grid lines."
        )

    except Exception as e:
        print(f"  ❌ LỖI kiểm tra Dashboard: {e}")
        failed += 1

    # ── 4. Kiểm tra Date Picker ──
    print("\n═══ KIỂM TRA DATE PICKER ═══")
    try:
        # Sheet _Dates KHÔNG CÒN tồn tại (đã chuyển sang Dashboard columns AA, AB)
        check(
            '_Dates' not in wb.sheetnames,
            "Sheet '_Dates' KHÔNG tồn tại (đúng như thiết kế mới).",
            "Sheet '_Dates' vẫn còn tồn tại!"
        )

        # Sheet _RateData tồn tại
        check(
            '_RateData' in wb.sheetnames,
            "Sheet '_RateData' tồn tại.",
            "Sheet '_RateData' KHÔNG tồn tại!"
        )

        if '_RateData' in wb.sheetnames:
            ws_rate = wb['_RateData']
            # Sheet _RateData ẩn
            check(
                ws_rate.sheet_state == 'hidden',
                "Sheet '_RateData' đã ẩn (hidden).",
                f"Sheet '_RateData' chưa ẩn (state: {ws_rate.sheet_state})."
            )
            # Header đầu tiên là 'Ngày'
            check(
                ws_rate.cell(1, 1).value == 'Ngày',
                "Sheet '_RateData' header đúng (cột A = 'Ngày').",
                f"Sheet '_RateData' header sai (cột A = '{ws_rate.cell(1, 1).value}')."
            )
            # Có dữ liệu
            check(
                ws_rate.max_row >= 2,
                f"Sheet '_RateData' có {ws_rate.max_row - 1} hàng dữ liệu.",
                "Sheet '_RateData' trống!"
            )

        # Data Validation trên ô B4 (ngày xem)
        ws_dash = wb['Dashboard']
        dv_found = False
        for dv in ws_dash.data_validations.dataValidation:
            for cell_range in dv.sqref.ranges:
                if 'B4' in str(cell_range):
                    dv_found = True
                    check(
                        dv.type == 'list',
                        f"Data Validation trên B4 là dropdown list (type={dv.type}).",
                        f"Data Validation trên B4 KHÔNG phải list (type={dv.type})."
                    )
                    break
        check(dv_found, "Ô B4 có Data Validation (date picker).", "Ô B4 KHÔNG có Data Validation!")

        # Kiểm tra formula INDEX/MATCH trên ô C7 (tên ngoại tệ đầu tiên)
        formula_cell = ws_dash['C7'].value
        check(
            formula_cell and 'INDEX' in str(formula_cell) and 'MATCH' in str(formula_cell),
            f"Ô C7 dùng formula INDEX/MATCH.",
            f"Ô C7 KHÔNG dùng INDEX/MATCH (value: {formula_cell})."
        )

    except Exception as e:
        print(f"  ❌ LỖI kiểm tra Date Picker: {e}")
        failed += 1

    # ── 5. Kiểm tra SO SÁNH 2 NGÀY ──
    print("\n═══ KIỂM TRA SO SÁNH 2 NGÀY ═══")
    try:
        ws_dash = wb['Dashboard']

        # Ô E4 có Data Validation (ngày so sánh)
        dv_compare_found = False
        for dv in ws_dash.data_validations.dataValidation:
            for cell_range in dv.sqref.ranges:
                if 'E4' in str(cell_range):
                    dv_compare_found = True
                    check(
                        dv.type == 'list',
                        f"Data Validation trên E4 là dropdown list.",
                        f"Data Validation trên E4 KHÔNG phải list."
                    )
                    break
        check(dv_compare_found, "Ô E4 có Data Validation (ngày so sánh).", "Ô E4 KHÔNG có Data Validation!")

        # Label "So sánh với ngày"
        check(
            ws_dash['E3'].value == 'So sánh với ngày',
            "Label E3 đúng: 'So sánh với ngày'.",
            f"Label E3 sai: '{ws_dash['E3'].value}'."
        )

        # Cột I header = "Bán (SS)"
        check(
            ws_dash['I6'].value == 'Bán (SS)',
            f"Header I6 đúng: '{ws_dash['I6'].value}'.",
            f"Header I6 sai: '{ws_dash['I6'].value}'."
        )

        # Cột I7 dùng formula INDEX/MATCH với $E$4
        i7_val = ws_dash['I7'].value
        check(
            i7_val and 'INDEX' in str(i7_val) and '$E$4' in str(i7_val),
            "Ô I7 dùng formula INDEX/MATCH trỏ đến $E$4.",
            f"Ô I7 formula sai (value: {i7_val})."
        )

        # Cột J header = "Chênh lệch"
        check(
            ws_dash['J6'].value == 'Chênh lệch',
            f"Header J6 đúng: '{ws_dash['J6'].value}'.",
            f"Header J6 sai: '{ws_dash['J6'].value}'."
        )

        # Cột K header = "% So sánh"
        check(
            ws_dash['K6'].value == '% So sánh',
            f"Header K6 đúng: '{ws_dash['K6'].value}'.",
            f"Header K6 sai: '{ws_dash['K6'].value}'."
        )

    except Exception as e:
        print(f"  ❌ LỖI kiểm tra so sánh 2 ngày: {e}")
        failed += 1

    # ── 6. Kiểm tra BIỂU ĐỒ VỚI CURRENCY SELECTOR ──
    print("\n═══ KIỂM TRA BIỂU ĐỒ & CURRENCY SELECTOR ═══")
    try:
        ws_dash = wb['Dashboard']

        # Sheet _ChartData tồn tại và ẩn
        check(
            '_ChartData' in wb.sheetnames,
            "Sheet '_ChartData' tồn tại.",
            "Sheet '_ChartData' KHÔNG tồn tại!"
        )

        if '_ChartData' in wb.sheetnames:
            ws_chart = wb['_ChartData']
            check(
                ws_chart.sheet_state == 'hidden',
                "Sheet '_ChartData' đã ẩn.",
                f"Sheet '_ChartData' chưa ẩn (state: {ws_chart.sheet_state})."
            )
            # Header đầu tiên = Tháng
            check(
                ws_chart.cell(1, 1).value == 'Tháng',
                "Sheet '_ChartData' header đúng (cột A = 'Tháng').",
                f"Sheet '_ChartData' header sai: '{ws_chart.cell(1, 1).value}'."
            )

        # Currency selector - tìm Data Validation loại list trỏ đến $AB
        curr_dv_found = False
        for dv in ws_dash.data_validations.dataValidation:
            if dv.formula1 and '$AB$1' in str(dv.formula1):
                curr_dv_found = True
                break
        check(curr_dv_found, "Có Data Validation cho currency selector (trỏ $AB).", "KHÔNG tìm thấy currency selector Data Validation!")

        # Dashboard sheet cột AB có ngoại tệ
        first_curr = ws_dash.cell(1, 28).value
        check(
            first_curr is not None and len(str(first_curr)) >= 3,
            f"Dashboard cột AB có ngoại tệ (đầu tiên: {first_curr}).",
            f"Dashboard cột AB trống hoặc sai (AB1: {first_curr})."
        )

        # Biểu đồ tồn tại trên Dashboard
        check(
            len(ws_dash._charts) > 0,
            f"Dashboard có {len(ws_dash._charts)} biểu đồ.",
            "Dashboard KHÔNG có biểu đồ!"
        )

    except Exception as e:
        print(f"  ❌ LỖI kiểm tra biểu đồ: {e}")
        failed += 1

    # ── 7. Kiểm tra CONDITIONAL FORMATTING ──
    print("\n═══ KIỂM TRA CONDITIONAL FORMATTING ═══")
    try:
        ws_dash = wb['Dashboard']
        cf_rules = ws_dash.conditional_formatting
        cf_count = len(list(cf_rules))

        check(
            cf_count > 0,
            f"Dashboard có {cf_count} conditional formatting rules.",
            "Dashboard KHÔNG có conditional formatting rules!"
        )

        # Kiểm tra có rule cho cột H
        has_h_rule = any('H' in str(rule.sqref) for rule in cf_rules)
        check(
            has_h_rule,
            "Có conditional formatting cho cột H (% Thay đổi).",
            "KHÔNG có conditional formatting cho cột H!"
        )

        # Kiểm tra có rule cho cột K
        has_k_rule = any('K' in str(rule.sqref) for rule in cf_rules)
        check(
            has_k_rule,
            "Có conditional formatting cho cột K (% So sánh).",
            "KHÔNG có conditional formatting cho cột K!"
        )

        # Kiểm tra có rule cho cột J
        has_j_rule = any('J' in str(rule.sqref) for rule in cf_rules)
        check(
            has_j_rule,
            "Có conditional formatting cho cột J (Chênh lệch).",
            "KHÔNG có conditional formatting cho cột J!"
        )

    except Exception as e:
        print(f"  ❌ LỖI kiểm tra conditional formatting: {e}")
        failed += 1

    # ── 8. Kiểm tra LOG FILE ──
    print("\n═══ KIỂM TRA LOG FILE ═══")
    check(
        os.path.exists(LOG_FILE),
        f"Log file tồn tại: {LOG_FILE}",
        f"Log file KHÔNG tồn tại: {LOG_FILE}"
    )

    if os.path.exists(LOG_FILE):
        log_size = os.path.getsize(LOG_FILE)
        check(
            log_size > 0,
            f"Log file có nội dung ({log_size} bytes).",
            "Log file trống!"
        )

    # ── 9. Kiểm tra CỘT BÌNH QUÂN VÀ CỘT % THAY ĐỔI ──
    print("\n═══ KIỂM TRA CỘT BÌNH QUÂN VÀ % THAY ĐỔI ═══")
    try:
        ws_dash = wb['Dashboard']

        # Header G6 = "Bình quân"
        check(
            ws_dash['G6'].value == 'Bình quân',
            f"Header G6 đúng: '{ws_dash['G6'].value}'.",
            f"Header G6 sai: '{ws_dash['G6'].value}'."
        )

        # G7 có formula AVERAGE
        g7_val = ws_dash['G7'].value
        check(
            g7_val and 'AVERAGE' in str(g7_val),
            f"Ô G7 dùng formula AVERAGE/IFERROR.",
            f"Ô G7 formula sai (value: {g7_val})."
        )

        # Header H6 = "% Thay đổi"
        check(
            ws_dash['H6'].value == '% Thay đổi',
            f"Header H6 đúng: '{ws_dash['H6'].value}'.",
            f"Header H6 sai: '{ws_dash['H6'].value}'."
        )

        # H7 có formula IFERROR
        h7_val = ws_dash['H7'].value
        check(
            h7_val and 'IFERROR' in str(h7_val),
            f"Ô H7 dùng formula IFERROR.",
            f"Ô H7 formula sai (value: {h7_val})."
        )

    except Exception as e:
        print(f"  ❌ LỖI kiểm tra cột % thay đổi: {e}")
        failed += 1

    # ── 10. Kiểm tra sheet TheoDoi_USD (So sánh ngân hàng) ──
    print("\n═══ KIỂM TRA SHEET THEODOI_USD ═══")
    try:
        check(
            'TheoDoi_USD' in wb.sheetnames,
            "Sheet 'TheoDoi_USD' đã được tạo thành công.",
            "Thiếu sheet 'TheoDoi_USD'!"
        )
        
        if 'TheoDoi_USD' in wb.sheetnames:
            ws_usd = wb['TheoDoi_USD']
            
            # Grid lines ẩn
            check(
                not ws_usd.sheet_view.showGridLines,
                "Sheet TheoDoi_USD ẩn grid lines.",
                "Sheet TheoDoi_USD chưa ẩn grid lines!"
            )
            
            # Tiêu đề
            check(
                ws_usd['B2'].value == 'BẢNG SO SÁNH TỶ GIÁ USD/VND GIỮA CÁC NGÂN HÀNG',
                "Tiêu đề sheet TheoDoi_USD đúng.",
                f"Tiêu đề sheet TheoDoi_USD sai: '{ws_usd['B2'].value}'"
            )
            
            # Các nhãn và bộ chọn
            check(
                ws_usd['B3'].value == 'Ngày xem' and ws_usd['E3'].value == 'So sánh với ngày',
                "Các nhãn ngày xem, so sánh với ngày hiển thị đúng.",
                f"Nhãn sai: B3='{ws_usd['B3'].value}', E3='{ws_usd['E3'].value}'"
            )
            
            # Headers bảng
            check(
                ws_usd['B7'].value == 'Ngân hàng' and ws_usd['C7'].value == 'Mua tiền mặt' and ws_usd['I7'].value == '% So sánh',
                "Headers bảng TheoDoi_USD đúng.",
                f"Headers sai: B7='{ws_usd['B7'].value}', C7='{ws_usd['C7'].value}', I7='{ws_usd['I7'].value}'"
            )
            
            # Dòng dữ liệu đầu tiên (dòng 8)
            check(
                'SUMIFS' in str(ws_usd['C8'].value) and 'Data_TheoDoi_USD' in str(ws_usd['C8'].value),
                "Các ô dữ liệu dòng 8 chứa đúng công thức động (SUMIFS).",
                f"Công thức sai: C8='{ws_usd['C8'].value}'"
            )
            
            # Data validation cho ô B4 và E4
            dv_found_b4 = False
            dv_found_e4 = False
            for dv in ws_usd.data_validations.dataValidation:
                for cell_range in dv.sqref.ranges:
                    if 'B4' in str(cell_range):
                        dv_found_b4 = True
                    if 'E4' in str(cell_range):
                        dv_found_e4 = True
            check(dv_found_b4 and dv_found_e4, "Các ô B4 và E4 đều có Data Validation.", f"Thiếu Validation: B4={dv_found_b4}, E4={dv_found_e4}")
            
    except Exception as e:
        print(f"  ❌ LỖI kiểm tra sheet TheoDoi_USD: {e}")
        failed += 1

    # ── 11. Kiểm tra sheet TheoDoi_Thang_USD (Báo cáo tháng) ──
    print("\n═══ KIỂM TRA SHEET THEODOI_THANG_USD ═══")
    try:
        check(
            'TheoDoi_Thang_USD' in wb.sheetnames,
            "Sheet 'TheoDoi_Thang_USD' đã được tạo thành công.",
            "Thiếu sheet 'TheoDoi_Thang_USD'!"
        )
        
        if 'TheoDoi_Thang_USD' in wb.sheetnames:
            ws_monthly = wb['TheoDoi_Thang_USD']
            
            # Grid lines ẩn
            check(
                not ws_monthly.sheet_view.showGridLines,
                "Sheet TheoDoi_Thang_USD ẩn grid lines.",
                "Sheet TheoDoi_Thang_USD chưa ẩn grid lines!"
            )
            
            # Tiêu đề
            check(
                ws_monthly['B2'].value == 'BÁO CÁO TỶ GIÁ USD THEO THÁNG',
                "Tiêu đề sheet TheoDoi_Thang_USD đúng.",
                f"Tiêu đề sheet TheoDoi_Thang_USD sai: '{ws_monthly['B2'].value}'"
            )
            
            # Các nhãn và bộ chọn
            check(
                ws_monthly['B4'].value == 'Chọn Năm:' and ws_monthly['E4'].value == 'Chọn Tháng:' and ws_monthly['G4'].value == 'Ngân Hàng:',
                "Các nhãn chọn Năm, chọn Tháng, chọn Ngân hàng hiển thị đúng.",
                f"Nhãn sai: B4='{ws_monthly['B4'].value}', E4='{ws_monthly['E4'].value}', G4='{ws_monthly['G4'].value}'"
            )
            
            # Headers bảng
            check(
                ws_monthly['C6'].value == 'Ngày theo lịch' and ws_monthly['D6'].value == 'Ngày áp dụng' and ws_monthly['H6'].value == 'Bình quân',
                "Headers bảng TheoDoi_Thang_USD đúng.",
                f"Headers sai: C6='{ws_monthly['C6'].value}', D6='{ws_monthly['D6'].value}', H6='{ws_monthly['H6'].value}'"
            )
            
            # Dòng dữ liệu đầu tiên (dòng 7)
            check(
                'DATE' in str(ws_monthly['C7'].value) and 'SUMIFS' in str(ws_monthly['E7'].value) and 'AVERAGE' in str(ws_monthly['H7'].value),
                "Các ô dữ liệu dòng 7 chứa đúng công thức động (DATE, SUMIFS, AVERAGE).",
                f"Công thức sai: C7='{ws_monthly['C7'].value}', E7='{ws_monthly['E7'].value}', H7='{ws_monthly['H7'].value}'"
            )
            
            # Data validation cho ô F4 và H4
            dv_month_found = False
            dv_bank_found = False
            for dv in ws_monthly.data_validations.dataValidation:
                for cell_range in dv.sqref.ranges:
                    if 'F4' in str(cell_range):
                        dv_month_found = True
                        check(
                            dv.type == 'list' and '$AA$1:$AA$12' in str(dv.formula1),
                            "Ô F4 có Data Validation dạng list trỏ tới danh sách tháng $AA$1:$AA$12.",
                            f"Validation F4 sai: type={dv.type}, formula={dv.formula1}"
                        )
                    if 'H4' in str(cell_range):
                        dv_bank_found = True
                        check(
                            dv.type == 'list' and '$AB$1:$AB$6' in str(dv.formula1),
                            "Ô H4 có Data Validation dạng list trỏ tới danh sách ngân hàng $AB$1:$AB$6.",
                            f"Validation H4 sai: type={dv.type}, formula={dv.formula1}"
                        )
            check(dv_month_found, "Ô F4 có Data Validation.", "Ô F4 KHÔNG có Data Validation!")
            check(dv_bank_found, "Ô H4 có Data Validation.", "Ô H4 KHÔNG có Data Validation!")
            
            # Dòng tính bình quân tháng (dòng 38)
            check(
                ws_monthly['C38'].value == 'BÌNH QUÂN THÁNG',
                "Dòng 38 nhãn đúng: 'BÌNH QUÂN THÁNG'.",
                f"Dòng 38 nhãn sai: '{ws_monthly['C38'].value}'"
            )
            check(
                'AVERAGE' in str(ws_monthly['E38'].value) and 'AVERAGE' in str(ws_monthly['H38'].value),
                "Các ô dòng 38 chứa công thức tính trung bình cả tháng.",
                f"Công thức trung bình sai: E38='{ws_monthly['E38'].value}', H38='{ws_monthly['H38'].value}'"
            )
            
    except Exception as e:
        print(f"  ❌ LỖI kiểm tra sheet TheoDoi_Thang_USD: {e}")
        failed += 1

    # ── Kết quả ──
    total = passed + failed
    print(f"\n{'═' * 40}")
    print(f"📊 KẾT QUẢ: {passed}/{total} passed")
    if failed == 0:
        print("🎉 TẤT CẢ KIỂM TRA ĐỀU PASS!")
    else:
        print(f"⚠  Có {failed} kiểm tra FAIL.")

if __name__ == "__main__":
    verify_excel()
