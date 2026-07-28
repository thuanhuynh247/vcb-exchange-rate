"""
Fetch VCB historical exchange rates from 2025-01-01 to present.
Merges with existing data, deduplicates, and creates a VCB-style Dashboard.
"""
import requests
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────────────
API_URL = 'https://www.vietcombank.com.vn/api/exchangerates'
OUTPUT_FILE = r'D:\Tygia-Tudong\TyGia_Banking.xlsx'
START_DATE = datetime(2025, 1, 1)
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                  'AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/131.0.0.0 Safari/537.36'
}

# Thứ tự hiển thị ngoại tệ chính (giống website VCB)
PRIORITY_CURRENCIES = ['USD', 'EUR', 'GBP', 'JPY', 'AUD', 'CAD', 'CHF', 'SGD',
                       'CNY', 'HKD', 'THB', 'KRW', 'MYR', 'INR', 'KWD',
                       'SAR', 'NOK', 'SEK', 'DKK', 'RUB']

# Ngoại tệ dùng cho biểu đồ Dashboard
CHART_CURRENCIES = ['USD', 'EUR']

# Mapping currency code -> English name
CURRENCY_NAMES = {
    'USD': 'US DOLLAR', 'EUR': 'EURO', 'GBP': 'POUND STERLING',
    'JPY': 'YEN', 'AUD': 'AUSTRALIAN DOLLAR', 'CAD': 'CANADIAN DOLLAR',
    'CHF': 'SWISS FRANC', 'SGD': 'SINGAPORE DOLLAR', 'CNY': 'YUAN RENMINBI',
    'HKD': 'HONGKONG DOLLAR', 'THB': 'THAILAND BAHT', 'KRW': 'KOREAN WON',
    'MYR': 'MALAYSIAN RINGGIT', 'INR': 'INDIAN RUPEE', 'KWD': 'KUWAITI DINAR',
    'SAR': 'SAUDI RIAL', 'NOK': 'NORWEGIAN KRONER', 'SEK': 'SWEDISH KRONA',
    'DKK': 'DANISH KRONE', 'RUB': 'RUSSIAN RUBLE',
}


# ── 1. Fetch historical rates ──────────────────────────────────────────
def fetch_rates_for_date(date_str: str) -> list[dict] | None:
    """Fetch exchange rates for a single date from VCB JSON API."""
    try:
        resp = requests.get(
            API_URL, params={'date': date_str},
            headers=HEADERS, timeout=15
        )
        resp.raise_for_status()
        data = resp.json()

        api_date = data.get('Date', '')[:10]  # 'YYYY-MM-DD'
        rows = []
        for item in data.get('Data', []):
            code = item.get('currencyCode', '')
            if code:  # Lấy tất cả ngoại tệ
                rows.append({
                    'Ngày Cập Nhật': api_date,
                    'Mã Ngoại Tệ': code,
                    'Tên Ngoại Tệ': CURRENCY_NAMES.get(code, item.get('currencyName', '').strip()),
                    'Mua Tiền Mặt': _to_float(item.get('cash', '')),
                    'Mua Chuyển Khoản': _to_float(item.get('transfer', '')),
                    'Bán': _to_float(item.get('sell', '')),
                })
        return rows if rows else None

    except Exception as e:
        print(f"  ⚠ Lỗi khi lấy dữ liệu ngày {date_str}: {e}")
        return None


def _to_float(val):
    try:
        v = float(str(val).replace(',', ''))
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None


def fetch_all_historical() -> pd.DataFrame:
    """Iterate day-by-day and collect exchange rate data."""
    end_date = datetime.now() - timedelta(days=1)  # yesterday

    current = START_DATE
    try:
        # if os.path.exists(OUTPUT_FILE):
        #     df_old = pd.read_excel(OUTPUT_FILE, sheet_name=0)
        #     if not df_old.empty and 'Ngày Cập Nhật' in df_old.columns:
        #         max_date_str = df_old['Ngày Cập Nhật'].max()
        #         if isinstance(max_date_str, str):
        #             max_date = datetime.strptime(max_date_str[:10], '%Y-%m-%d')
        #         else:
        #             max_date = pd.to_datetime(max_date_str)
        #         if max_date >= START_DATE:
        #             current = max_date + timedelta(days=1)
        #             print(f"📊 Tìm thấy dữ liệu cũ đến ngày {max_date.date()}. Sẽ tiếp tục từ {current.date()}")
        pass
    except Exception as e:
        print(f"⚠ Không thể đọc file cũ để xác định ngày bắt đầu: {e}")

    all_rows = []
    total_days = (end_date - current).days + 1

    print(f"📡 Bắt đầu lấy dữ liệu từ {current.date()} → {end_date.date()} ({total_days} ngày)")
    print("=" * 60)

    import concurrent.futures
    dates_to_fetch = []
    curr = current
    while curr <= end_date:
        dates_to_fetch.append(curr.strftime('%Y-%m-%d'))
        curr += timedelta(days=1)

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        future_to_date = {executor.submit(fetch_rates_for_date, d): d for d in dates_to_fetch}
        for i, future in enumerate(concurrent.futures.as_completed(future_to_date)):
            try:
                rows = future.result()
                if rows:
                    all_rows.extend(rows)
                if (i + 1) % 30 == 0 or (i + 1) == len(dates_to_fetch):
                    print(f"  ✅ Đã tải xong {i + 1}/{len(dates_to_fetch)} ngày")
            except Exception as exc:
                print(f"⚠ Lỗi tải: {exc}")

    print(f"\n📊 Hoàn tất: thu thập được {len(all_rows)} bản ghi từ API.")
    return pd.DataFrame(all_rows)


# ── 2. Merge & deduplicate ─────────────────────────────────────────────
def merge_with_existing(df_new: pd.DataFrame) -> pd.DataFrame:
    """Merge new historical data with existing Excel data, deduplicate."""
    if os.path.exists(OUTPUT_FILE):
        try:
            df_old = pd.read_excel(OUTPUT_FILE, sheet_name=0)
            print(f"📂 Đọc file hiện có: {len(df_old)} dòng.")
            df_old['_date_key'] = pd.to_datetime(
                df_old['Ngày Cập Nhật']
            ).dt.strftime('%Y-%m-%d')
        except Exception as e:
            print(f"  ⚠ Không đọc được file cũ, dùng dữ liệu mới: {e}")
            df_old = pd.DataFrame()
    else:
        df_old = pd.DataFrame()

    if not df_new.empty:
        df_new['_date_key'] = df_new['Ngày Cập Nhật']

    if not df_old.empty and not df_new.empty:
        df_combined = pd.concat([df_new, df_old], ignore_index=True)
    elif not df_new.empty:
        df_combined = df_new.copy()
    else:
        df_combined = df_old.copy()

    if df_combined.empty:
        return df_combined

    # Dedup: keep the first occurrence per (date, currency)
    df_combined = df_combined.drop_duplicates(
        subset=['_date_key', 'Mã Ngoại Tệ'], keep='first'
    )
    df_combined = df_combined.sort_values(
        ['_date_key', 'Mã Ngoại Tệ'], ascending=[False, True]
    ).reset_index(drop=True)

    df_combined['Ngày Cập Nhật'] = df_combined['_date_key']
    df_combined.drop(columns=['_date_key'], inplace=True)

    # Đảm bảo có cột Tên Ngoại Tệ
    if 'Tên Ngoại Tệ' not in df_combined.columns:
        df_combined.insert(2, 'Tên Ngoại Tệ', df_combined['Mã Ngoại Tệ'].map(CURRENCY_NAMES).fillna(''))

    print(f"✅ Sau dedup: {len(df_combined)} bản ghi.")
    return df_combined


# ── 3. Write Excel with formatting ────────────────────────────────────
def write_excel(df: pd.DataFrame):
    """Write data to 'Data' sheet + create 'Dashboard' sheet with date picker."""
    if df.empty:
        print("❌ Không có dữ liệu để ghi.")
        return

    # Import dashboard function from get_rates.py (same directory)
    import sys
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if script_dir not in sys.path:
        sys.path.insert(0, script_dir)
    from get_rates import _format_data_sheet, _create_dashboard_with_date_picker

    print("Bước 1: Lưu Data sheet...")
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data', index=False)

    print("Bước 2: Format Data sheet...")
    wb = load_workbook(OUTPUT_FILE)
    ws_data = wb['Data']
    _format_data_sheet(ws_data)
    wb.save(OUTPUT_FILE)
    
    print("Bước 3: Khởi tạo Dashboard...")
    try:
        wb = load_workbook(OUTPUT_FILE)
        _create_dashboard_with_date_picker(wb, df)
        wb.save(OUTPUT_FILE)
        print(f"💾 Đã lưu thành công Dashboard với {len(df)} dòng dữ liệu.")
    except Exception as e:
        print(f"❌ Lỗi khi khởi tạo Dashboard: {e}")
        import traceback
        traceback.print_exc()
        print("Bản ghi dữ liệu (Data sheet) vẫn được giữ an toàn.")


# ── Main ───────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("🚀 VCB Historical Exchange Rate Fetcher")
    print("=" * 60)

    # Step 1: Fetch historical data
    df_hist = fetch_all_historical()

    # Step 2: Merge with existing
    df_final = merge_with_existing(df_hist)

    # Step 3: Write Excel + Dashboard
    write_excel(df_final)

    print("\n✨ Hoàn tất!")



