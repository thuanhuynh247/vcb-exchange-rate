import openpyxl
wb = openpyxl.load_workbook(r"D:\Tygia-Tudong\TyGia_Banking.xlsx", data_only=True)
print("Sheet names:", wb.sheetnames)
for name in wb.sheetnames:
    if not name.startswith('_') and not name.startswith('Data_'):
        ws = wb[name]
        print(f"\nSheet: {name}")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
        for r in range(1, min(10, ws.max_row + 1)):
            row_vals = [ws.cell(r, c).value for c in range(1, min(12, ws.max_column + 1))]
            print(f"  Row {r}: {row_vals}")
